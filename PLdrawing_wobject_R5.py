import tkinter
from tkinter import ttk
from PIL import ImageTk, Image
import openpyxl
import math
import random
import sys
sys.path.append('C:/Users/craig/Dropbox/00 - Cloud Documents/06 - Python Scripts')
import CraigsFunFunctions
import numpy
import tkinter.font

BED_IMAGE_FILEPATH = "PDFXEdit_2020-08-25_19-36-01.png"
TOP_IMAGE_FILEPATH = "PDFXEdit_2020-08-25_19-36-01.png"
XL_FILEPATH = 'B56 - VRF_Rooftop Amenity Investigation - 2020.08.17.xlsm'

class FuncVars(object):
    def __init__(self, parent):
        self.parent = parent
        #open workbook
        wb = openpyxl.load_workbook(XL_FILEPATH, data_only=True)
        ws = wb['Input LwA_XYZ']

        #initialize eqmt list
        self.equipment_list = list()
        for count, eqmt_tag, path, make, model, sound_level, sound_ref_dist, x_coord, y_coord, z_coord in zip(ws['A'], ws['B'], ws['C'], ws['D'], ws['E'], ws['F'], ws['G'], ws['J'], ws['K'], ws['L'] ):
            if count.value == "Number of Units": continue
            if count.value == None: break
            self.equipment_list.append(Equipment(count.value, str(eqmt_tag.value), path.value, make.value, model.value, sound_level.value, sound_ref_dist.value, x_coord.value, y_coord.value, z_coord.value))

        #initialize rcvr list
        self.receiver_list = list()
        for r_name, x_coord, y_coord, z_coord, sound_limit in zip(ws['P'], ws['Q'], ws['R'], ws['S'], ws['T']):
            if r_name.value == "R#": continue
            if r_name.value == None: break
            self.receiver_list.append(Receiver(str(r_name.value), x_coord.value, y_coord.value, z_coord.value, sound_limit.value, "NA"))

        #initialize barrier list
        self.barrier_list = list()
        for barrier_name, x0_coord, y0_coord, z0_coord, x1_coord, y1_coord, z1_coord in zip(ws['P'], ws['Q'], ws['R'], ws['S'], ws['T'], ws['U'], ws['V']):
            if int(barrier_name.coordinate[1:]) < 24: continue
            if barrier_name.value == None: break
            self.barrier_list.append(Barrier(str(barrier_name.value), x0_coord.value, y0_coord.value, z0_coord.value, x1_coord.value, y1_coord.value, z1_coord.value))

        #initialize master_scale
        self.old_master_scale = 1.0
        self.known_distance_ft = ws['U20'].value if ws['U20'].value != None else 1.0
        self.scale_line_distance_px = ws['V20'].value if ws['V20'].value != None else 1.0
        self.master_scale = self.known_distance_ft / self.scale_line_distance_px

    def update_master_scale(self, scale_line_distance_px, known_distance_ft):
        self.scale_line_distance_px = scale_line_distance_px
        self.known_distance_ft = known_distance_ft
        self.old_master_scale = self.master_scale
        self.master_scale = self.known_distance_ft / self.scale_line_distance_px

        '''rescaling eqmt'''
        for obj in self.equipment_list:
            obj.x_coord /= self.old_master_scale
            obj.y_coord /= self.old_master_scale
            obj.x_coord *= self.master_scale
            obj.y_coord *= self.master_scale
            obj.x_coord = round(obj.x_coord, 2)
            obj.y_coord = round(obj.y_coord, 2)

        for obj in self.receiver_list:
            obj.x_coord /= self.old_master_scale
            obj.y_coord /= self.old_master_scale
            obj.x_coord *= self.master_scale
            obj.y_coord *= self.master_scale
            obj.x_coord = round(obj.x_coord, 2)
            obj.y_coord = round(obj.y_coord, 2)
        '''rescaling eqmt'''

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateRcvrTree()
        self.parent.pane_eqmt_info.generateEqmtTree()

class Equipment(object):
    def __init__(self, count, eqmt_tag, path, make, model, sound_level, sound_ref_dist, x_coord, y_coord, z_coord):
        self.count = count
        self.eqmt_tag = eqmt_tag.replace(" ", "-")
        self.path = path
        self.make = make
        self.model = model
        self.sound_level = sound_level if sound_level != None else 0
        self.sound_ref_dist = sound_ref_dist if sound_ref_dist != None else 0
        self.x_coord = x_coord if x_coord != None else 0
        self.y_coord = y_coord if y_coord != None else 0
        self.z_coord = z_coord if z_coord != None else 0

class Receiver(object):
    def __init__(self, r_name, x_coord, y_coord, z_coord, sound_limit, predicted_sound_level):
        self.r_name = r_name.replace(" ", "-")
        self.x_coord = x_coord if x_coord != None else 0
        self.y_coord = y_coord if y_coord != None else 0
        self.z_coord = z_coord if z_coord != None else 0
        self.sound_limit = sound_limit
        self.predicted_sound_level = predicted_sound_level

class Barrier(object):
    def __init__(self, barrier_name, x0_coord, y0_coord, z0_coord, x1_coord, y1_coord, z1_coord):
        self.barrier_name = barrier_name.replace(" ", "-")
        self.x0_coord = x0_coord if x0_coord != None else 0
        self.y0_coord = y0_coord if y0_coord != None else 0
        self.z0_coord = z0_coord if z0_coord != None else 0
        self.x1_coord = x1_coord if x1_coord != None else 0
        self.y1_coord = y1_coord if y1_coord != None else 0
        self.z1_coord = z1_coord if z1_coord != None else 0

class Editor(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, parent)
        self.parent = parent

        #open image
        self.image = Image.open(BED_IMAGE_FILEPATH)
        self.image2 = Image.open(TOP_IMAGE_FILEPATH)

        #image sizing
        self.imageWidth, self.imageHeight = self.image.size
        print(self.image.size)
        self.image_size_factor = 1.5
        self.imageWidth *= self.image_size_factor
        self.imageHeight *= self.image_size_factor
        self.imageWidth = int(self.imageWidth)
        self.imageHeight = int(self.imageHeight)
        self.image = self.image.resize((self.imageWidth, self.imageHeight), Image.LANCZOS)
        self.image2_new_width = int(self.image2.size[0]/2)
        self.image2_new_height = int(self.image2.size[1]/2)
        self.image2 = self.image2.resize((self.image2_new_width, self.image2_new_height), Image.LANCZOS)
        self.tk_image = ImageTk.PhotoImage(self.image)
        self.tk_image2 = ImageTk.PhotoImage(self.image2)
        #self.tk_image2 = ImageTk.PhotoImage(self.image2.rotate(100))

        #canvas sizing
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.canvas_size_factor = 1
        self.canvasWidth = self.screen_width * self.canvas_size_factor
        self.canvasHeight = self.screen_height * self.canvas_size_factor
        self.canvasWidth -= 1000 # otherwise window is off the screen on home pc
        self.canvasHeight -= 250 # otherwise window is off the screen on home pc
        self.canvas = tkinter.Canvas(self, width=self.canvasWidth , height=self.canvasHeight, cursor="cross")

        self.canvas.config(scrollregion=(0, 0, self.imageWidth, self.imageHeight)) #giving scrollbars
        self.canvas.create_image(0,0, anchor="nw", image=self.tk_image, tag="bed_layer")
        image2_x_coord = self.image2.size[0]/2
        image2_y_coord = self.image2.size[1]/2
        self.canvas.create_image(image2_x_coord, image2_y_coord, tag="eqmt_drawing", image=self.tk_image2)

        '''scroll bar setup'''
        self.vScrollbar = tkinter.Scrollbar(self, orient=tkinter.VERTICAL)
        self.hScrollbar = tkinter.Scrollbar(self, orient=tkinter.HORIZONTAL)
        self.vScrollbar.config(command=self.canvas.yview)
        self.hScrollbar.config(command=self.canvas.xview)
        self.canvas.config(yscrollcommand=self.vScrollbar.set)
        self.canvas.config(xscrollcommand=self.hScrollbar.set)

        self.canvas.grid(row=0, column=0, sticky=tkinter.N + tkinter.S + tkinter.E + tkinter.W)
        self.vScrollbar.grid(row=0, column=1, stick=tkinter.N + tkinter.S)
        self.hScrollbar.grid(row=1, column=0, sticky=tkinter.E + tkinter.W)
        '''scroll bar setup'''

        '''initialize receivers and equipment boxes'''
        for eqmt in self.parent.func_vars.equipment_list:
            random_8bit_color = CraigsFunFunctions.random_8bit_color()
            offset = 20
            x = eqmt.x_coord/self.parent.func_vars.master_scale
            y = eqmt.y_coord/self.parent.func_vars.master_scale
            # self.canvas.coords(self.temp_rect, self.x0-10, self.y0-10, self.curX+10, self.curY+10)
            self.rectPerm = self.canvas.create_rectangle(x-offset, y-offset, x+offset, y+offset, tag=eqmt.eqmt_tag, fill=random_8bit_color, activeoutline='red')
            self.canvas.create_text(x, y, tag=eqmt.eqmt_tag, text=eqmt.eqmt_tag, font=("arial.ttf", 15), fill='Black')

        for rcvr in self.parent.func_vars.receiver_list:
            random_8bit_color = CraigsFunFunctions.random_8bit_color()
            offset = 20
            x = rcvr.x_coord/self.parent.func_vars.master_scale
            y = rcvr.y_coord/self.parent.func_vars.master_scale
            # self.canvas.coords(self.temp_rect, self.x0-10, self.y0-10, self.curX+10, self.curY+10)
            self.rectPerm = self.canvas.create_rectangle(x-offset, y-offset, x+offset, y+offset, tag=rcvr.r_name, fill=random_8bit_color, activeoutline='red')
            self.canvas.create_text(x, y, tag=rcvr.r_name, text=rcvr.r_name, font=("arial.ttf", 15), fill='Black')
        '''initialize receivers and equipment boxes'''

        self.temp_rect = None
        self.scale_line = None
        self.measure_line = None
        self.angle = 0

        self.canvas.bind("<Shift-ButtonPress-1>", self.shift_click)
        self.canvas.bind("<Shift-B1-Motion>", self.shift_click_move)
        self.canvas.bind("<Shift-ButtonRelease-1>", self.shift_click_release)

        '''Scrollable image'''
        self.canvas.bind('<Enter>', self._bound_to_mousewheel)
        self.canvas.bind('<Leave>', self._unbound_to_mousewheel)

    def _bound_to_mousewheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Shift-MouseWheel>", self._on_shift_mousewheel)
    def _unbound_to_mousewheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")
    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    def _on_shift_mousewheel(self, event):
        self.canvas.xview_scroll(int(-1*(event.delta/120)), "units")
        '''Scrollable image'''

    def get_angle(self, x, y):
        v0 = [x, 0]
        v1 = [x, y]
        dot_product = numpy.dot(v0, v1)
        v0_mag = numpy.linalg.norm(v0)
        v1_mag = numpy.linalg.norm(v1)
        angle = math.degrees(numpy.arccos((dot_product/(v0_mag*v1_mag))))
        if dot_product < 0:
            angle += 90
        print("hey", angle)
        return angle
    def update_distance_label(self):
        self.parent.pane_eqmt_info.measuremet_label.configure(text=str(round(self.parent.func_vars.master_scale*(math.sqrt((self.x0 - self.curX)**2 + (self.y0 - self.curY)**2)),2)) + " ft")
    def get_current_n_start_mouse_pos(self, event):
        self.x0 = self.canvas.canvasx(event.x)
        self.y0 = self.canvas.canvasy(event.y)
        self.curX = self.canvas.canvasx(event.x)
        self.curY = self.canvas.canvasy(event.y)
    def get_current_mouse_pos(self, event):
        self.curX = self.canvas.canvasx(event.x)
        self.curY = self.canvas.canvasy(event.y)

    def setting_scale_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)

        if self.scale_line != None:
            self.canvas.delete(self.scale_line)
        self.temp_scale_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="orange", width=5)
    def setting_scale_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_scale_line, self.x0, self.y0, self.curX, self.curY)
    def setting_scale_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_scale_line)

        self.scale_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="blue", width=5)
        scale_line_coords = self.canvas.coords(self.scale_line)
        # self.parent.func_vars.scale_line_distance_px = CraigsFunFunctions.distance_formula(scale_line_coords[0], scale_line_coords[2], scale_line_coords[1], scale_line_coords[3])
        # self.parent.func_vars.known_distance_ft = float(self.parent.pane_eqmt_info.e1.get())
        _scale_line_distance_px = CraigsFunFunctions.distance_formula(scale_line_coords[0], scale_line_coords[2], scale_line_coords[1], scale_line_coords[3])
        _known_distance_ft = float(self.parent.pane_eqmt_info.e1.get())
        self.parent.func_vars.update_master_scale(_scale_line_distance_px, _known_distance_ft)
        # self.parent.func_vars.old_master_scale = self.parent.func_vars.master_scale
        # self.parent.func_vars.master_scale = self.parent.func_vars.known_distance_ft / self.parent.func_vars.scale_line_distance_px

        scaleIndicatorLabelText = "Scale: " + str(round(self.parent.func_vars.scale_line_distance_px,0)) + " px = " + str(self.parent.func_vars.known_distance_ft) + " ft"
        self.parent.pane_eqmt_info.scaleIndicatorLabel.configure(text=scaleIndicatorLabelText)

    def drawing_eqmt_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.temp_rect = self.canvas.create_rectangle(self.x0, self.y0, self.x0, self.y0, outline='red')
    def drawing_eqmt_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_rect, self.x0, self.y0, self.curX, self.curY)
    def drawing_eqmt_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_rect)

        random_8bit_color = CraigsFunFunctions.random_8bit_color()

        eqmt_tag=self.parent.pane_eqmt_info.current_euqipment[1] # i think this is grabbing from the tree
        tagged_objects = self.canvas.find_withtag(eqmt_tag)
        for tagged_object in tagged_objects:
            self.canvas.delete(tagged_object)
        self.rectPerm = self.canvas.create_rectangle(self.x0, self.y0, self.curX, self.curY, tag=eqmt_tag, fill=random_8bit_color, activeoutline='red')

        self.canvas.create_text((self.x0 + (self.curX-self.x0)/2, self.y0 + (self.curY - self.y0)/2), tag=eqmt_tag, text=eqmt_tag, font=("arial.ttf", 15), fill='Black')

        #update this one piece of eqmt
        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == eqmt_tag:
                obj.x_coord = self.x0 + (self.curX - self.x0)/2
                obj.y_coord = self.y0 + (self.curY - self.y0)/2
                obj.x_coord *= self.parent.func_vars.master_scale
                obj.y_coord *= self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)
                # print(obj.x_coord)
                # print(obj.y_coord)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateRcvrTree()
        self.parent.pane_eqmt_info.generateEqmtTree()

    def drawing_rcvr_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.temp_rect = self.canvas.create_rectangle(self.x0, self.y0, self.x0, self.y0, outline='green')
    def drawing_rcvr_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_rect, self.x0, self.y0, self.curX, self.curY)
    def drawing_rcvr_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_rect)

        random_8bit_color = CraigsFunFunctions.random_8bit_color()

        r_name=self.parent.pane_eqmt_info.current_receiver[0]
        tagged_objects = self.canvas.find_withtag(r_name)
        for tagged_object in tagged_objects:
            self.canvas.delete(tagged_object)
        self.rectPerm = self.canvas.create_rectangle(self.x0, self.y0, self.curX, self.curY, tag=r_name, fill=random_8bit_color, activeoutline='red')

        self.canvas.create_text((self.x0 + (self.curX-self.x0)/2, self.y0 + (self.curY - self.y0)/2), tag=r_name, text=r_name, font=("arial.ttf", 15), fill='Black')

        #update this one rcvr
        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == r_name:
                obj.x_coord = self.x0 + (self.curX - self.x0)/2
                obj.y_coord = self.y0 + (self.curY - self.y0)/2
                obj.x_coord *= self.parent.func_vars.master_scale
                obj.y_coord *= self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateRcvrTree()

    def measureing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        if self.measure_line != None:
            self.canvas.delete(self.measure_line)
        self.update_distance_label()
        self.temp_measure_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="orange", width=5)
    def measureing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_measure_line, self.x0, self.y0, self.curX, self.curY)
        self.update_distance_label()
    def measureing_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_measure_line)
        self.measure_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="red", width=5)

    def rotating_eqmt_drawing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        #calc angle at start point
        self.eqmt_drawing_center = self.canvas.coords("eqmt_drawing")
        self.angle0 = self.get_angle(self.x0-self.eqmt_drawing_center[0], self.y0-self.eqmt_drawing_center[1]) - self.angle
    def rotating_eqmt_drawing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        # calculate current angle relative to initial angle
        self.angle_1 = self.get_angle(self.curX-self.eqmt_drawing_center[0], self.curY-self.eqmt_drawing_center[1])
        dwg_x = self.eqmt_drawing_center[0]
        dwg_y = self.eqmt_drawing_center[1]

        if self.curX > dwg_x and self.curY > dwg_y:
            self.angle_1 *= -1
        elif self.curX < dwg_x and self.curY < dwg_y:
            self.angle_1 = 180 - self.angle_1
        elif self.curX < dwg_x and self.curY > dwg_y:
            self.angle_1 = self.angle_1 + 180
        elif self.curX < dwg_x and self.curY > dwg_y:
            self.angle_1 = self.angle_1 + 360

        self.angle = self.angle_1

        self.canvas.delete("eqmt_drawing")
        self.tk_image2 = ImageTk.PhotoImage(self.image2.rotate(self.angle, expand=True))
        self.canvas.create_image(self.eqmt_drawing_center[0], self.eqmt_drawing_center[1], image=self.tk_image2, tag="eqmt_drawing")
        self.canvas.tag_lower("eqmt_rdawing")
        self.canvas.tag_lower("bed_layer")

    def moving_eqmt_drawing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.eqmt_drawing_center = self.canvas.coords("eqmt_drawing")
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.delete("eqmt_drawing")
        self.canvas.create_image(self.eqmt_drawing_center[0] + x_shifter, self.eqmt_drawing_center[1] + y_shifter, image=self.tk_image2, tag="eqmt_drawing")
    def moving_eqmt_drawing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        # self.eqmt_drawing_center = self.canvas.coords("eqmt_drawing")
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.delete("eqmt_drawing")
        self.canvas.create_image(self.eqmt_drawing_center[0] + x_shifter, self.eqmt_drawing_center[1] + y_shifter, image=self.tk_image2, tag="eqmt_drawing")
        self.canvas.tag_lower("eqmt_drawing")
        self.canvas.tag_lower("bed_layer")

    def resizing_eqmt_drawing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.eqmt_dwg_cntr = self.canvas.coords("eqmt_drawing")
        self.eqmt_dwg_width_0 = self.image2_new_width
        self.eqmt_dwg_height_0 = self.image2_new_height
        self.eqmt_dwg_ratio = self.eqmt_dwg_width_0 / self.eqmt_dwg_height_0
        self.rect_p1_x0 = self.eqmt_dwg_cntr[0]-self.eqmt_dwg_width_0/2
        self.rect_p2_x0 = self.eqmt_dwg_cntr[0]+self.eqmt_dwg_width_0/2
        self.rect_p1_y0 = self.eqmt_dwg_cntr[1]-self.eqmt_dwg_height_0/2
        self.rect_p2_y0 = self.eqmt_dwg_cntr[1]+self.eqmt_dwg_height_0/2
        self.temp_rect = self.canvas.create_rectangle(self.rect_p1_x0, self.rect_p1_y0, self.rect_p2_x0, self.rect_p2_y0, outline='red')
    def resizing_eqmt_drawing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.x_change = self.curX - self.x0
        self.y_change = self.curY - self.y0
        self.rect_p1_x1 = self.rect_p1_x0
        self.rect_p2_x2 = self.rect_p2_x0 + self.x_change
        self.rect_p1_y1 = self.rect_p1_y0
        self.rect_p2_y1 = self.rect_p2_y0 + self.x_change / self.eqmt_dwg_ratio
        self.canvas.coords(self.temp_rect, self.rect_p1_x1, self.rect_p1_y1, self.rect_p2_x2, self.rect_p2_y1)
    def resizing_eqmt_drawing_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.eqmt_dwg_width_1 = self.eqmt_dwg_width_0 + int(self.x_change)
        self.eqmt_dwg_height_1 = int(self.eqmt_dwg_width_1 / self.eqmt_dwg_ratio)
        self.eqmt_dwg_cntr[0] += int(self.x_change/2)
        self.eqmt_dwg_cntr[1] += int((self.x_change / self.eqmt_dwg_ratio)/2)

        self.canvas.delete("eqmt_drawing")
        self.canvas.delete(self.temp_rect)
        self.image2 = self.image2.resize((self.eqmt_dwg_width_1, self.eqmt_dwg_height_1), Image.LANCZOS)
        self.tk_image2 = ImageTk.PhotoImage(self.image2.rotate(self.angle, expand=True))
        self.canvas.create_image(self.eqmt_dwg_cntr[0], self.eqmt_dwg_cntr[1], image=self.tk_image2, tag="eqmt_drawing")

        self.image2_new_width = self.eqmt_dwg_width_1
        self.image2_new_height = self.eqmt_dwg_height_1

        self.canvas.tag_lower("eqmt_drawing")
        self.canvas.tag_lower("bed_layer")


    def shift_click(self, event):
        if self.canvas.find_withtag("current"):
            self.eqmt_rcvr_tagged = self.canvas.gettags("current")
            self.tag_or_rcvr_num = self.eqmt_rcvr_tagged[0]
            self.eqmt_rcvr_ids = self.canvas.find_withtag(self.eqmt_rcvr_tagged[0])
            self.current_rect = self.eqmt_rcvr_ids[0]
            self.current_text = self.eqmt_rcvr_ids[1]
            self.current_rect_coords = self.canvas.coords(self.current_rect)
            self.current_text_coords = self.canvas.coords(self.current_text)

            self.get_current_n_start_mouse_pos(event)

        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == self.tag_or_rcvr_num:
                self.obj_x_coord_0 = obj.x_coord
                self.obj_y_coord_0 = obj.y_coord

        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == self.tag_or_rcvr_num:
                self.obj_x_coord_0 = obj.x_coord
                self.obj_y_coord_0 = obj.y_coord
    def shift_click_move(self, event):
        self.get_current_mouse_pos(event)
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.coords(self.current_rect, self.current_rect_coords[0]+x_shifter, self.current_rect_coords[1]+y_shifter, self.current_rect_coords[2]+x_shifter, self.current_rect_coords[3]+y_shifter)
        self.canvas.coords(self.current_text, self.current_text_coords[0]+x_shifter, self.current_text_coords[1]+y_shifter)

        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == self.tag_or_rcvr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == self.tag_or_rcvr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateEqmtTree()
        self.parent.pane_eqmt_info.generateRcvrTree()
    def shift_click_release(self, event):
        self.get_current_mouse_pos(event)
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.coords(self.current_rect, self.current_rect_coords[0]+x_shifter, self.current_rect_coords[1]+y_shifter, self.current_rect_coords[2]+x_shifter, self.current_rect_coords[3]+y_shifter)
        self.canvas.coords(self.current_text, self.current_text_coords[0]+x_shifter, self.current_text_coords[1]+y_shifter)

        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == self.tag_or_rcvr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == self.tag_or_rcvr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateEqmtTree()
        self.parent.pane_eqmt_info.generateRcvrTree()

class Pane_Toolbox(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, parent)
        self.parent = parent

        self.button_set_image_scale = tkinter.Button(self, text="Set Image Scale", command=self.set_scale, font=(None, 15))
        self.button_draw_equipment = tkinter.Button(self, text="Draw Equipment", command=self.draw_equipment, font=(None, 15))
        self.button_draw_receiver = tkinter.Button(self, text="Draw Receiver", command=self.draw_receiver, font=(None, 15))
        self.button_measure = tkinter.Button(self, text="Measure", command=self.measure, font=(None, 15))
        self.button_rotate_eqmt_drawing = tkinter.Button(self, text="Rotate Eqmt Drawing", command=self.rotate_eqmt_drawing, font=(None, 15))
        self.button_move_eqmt_drawing = tkinter.Button(self, text="Move Eqmt Drawing", command=self.move_eqmt_drawing, font=(None, 15))
        self.button_resize_eqmt_drawing = tkinter.Button(self, text="Resize Eqmt Drawing", command=self.resize_eqmt_drawing, font=(None, 15))

        self.button_set_image_scale.grid(row=0, column=0, sticky=tkinter.N + tkinter.W)
        self.button_draw_equipment.grid(row=1, column=0, sticky=tkinter.N + tkinter.W)
        self.button_draw_receiver.grid(row=2, column=0, sticky=tkinter.N + tkinter.W)
        self.button_measure.grid(row=2, column=0, sticky=tkinter.N + tkinter.W)
        self.button_rotate_eqmt_drawing.grid(row=3, column=0, sticky=tkinter.N + tkinter.W)
        self.button_move_eqmt_drawing.grid(row=4, column=0, sticky=tkinter.N + tkinter.W)
        self.button_resize_eqmt_drawing.grid(row=5, column=0, sticky=tkinter.N + tkinter.W)

    def set_scale(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.setting_scale_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.setting_scale_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.setting_scale_leftMouseRelease)

        self.parent.pane_eqmt_info.status_label.configure(text='Status: Setting Scale')
        self.parent.pane_eqmt_info.e1.focus()

    def draw_equipment(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.drawing_eqmt_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.drawing_eqmt_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.drawing_eqmt_leftMouseRelease)

        self.parent.pane_eqmt_info.status_label.configure(text='Status: Drawing Equipment')

    def draw_receiver(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.drawing_rcvr_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.drawing_rcvr_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.drawing_rcvr_leftMouseRelease)
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Drawing Receiver')

    def measure(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.measureing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.measureing_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.measureing_leftMouseRelease)
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Measuring')

    def rotate_eqmt_drawing(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.rotating_eqmt_drawing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.rotating_eqmt_drawing_leftMouseMove)
        self.parent.editor.canvas.unbind("<ButtonRelease-1>")
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Rotating Equipment Drawing')

    def move_eqmt_drawing(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.moving_eqmt_drawing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.moving_eqmt_drawing_leftMouseMove)
        self.parent.editor.canvas.unbind("<ButtonRelease-1>")
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Moving Equipment Drawing')

    def resize_eqmt_drawing(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.resizing_eqmt_drawing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.resizing_eqmt_drawing_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.resizing_eqmt_drawing_leftMouseRelease)
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Resizing Equipment Drawing')

class Pane_Eqmt_Info(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, parent)
        self.parent = parent
        self.update_est_noise_levels()
        
        self.myFont = tkinter.font.nametofont('TkTextFont')

        self.e1 = tkinter.Entry(self, font=(None, 15), width=36)
        self.e1.insert(0, "input scale & eqmt_tag names here prior to setting")
        self.e1.bind("<FocusIn>", self.e1_select_all)
        self.e1.bind("<Return>", self.e1_unfocus)

        scaleIndicatorLabelText = "Scale: " + str(round(self.parent.func_vars.scale_line_distance_px,0)) + " px = " + str(self.parent.func_vars.known_distance_ft) + " ft"

        self.exportList_button = tkinter.Button(self, text="Export Tag List", command=self.onExportListButton, font=(None, 15))
        self.scaleIndicatorLabel = tkinter.Label(self, text=scaleIndicatorLabelText, borderwidth=2, relief="solid", font=(None, 15))
        self.equipment_list_label = tkinter.Label(self, text="Equipment", font=(None, 15))
        self.status_label = tkinter.Label(self, text="Status: Idle", borderwidth=2, relief="solid", font=(None, 15))
        self.measuremet_label = tkinter.Label(self, text="", borderwidth=2, relief="solid", font=(None, 15))
        self.receiver_list_label = tkinter.Label(self, text="Receivers", font=(None, 15))
        self.generateEqmtTree()
        self.generateRcvrTree()
        self.generateBarrierTree()

        self.e1.grid(row=0, column=1, sticky=tkinter.N + tkinter.W)
        self.exportList_button.grid(row=1, column=1, sticky=tkinter.N + tkinter.W)
        self.scaleIndicatorLabel.grid(row=2, column=1, sticky=tkinter.N + tkinter.W)
        self.status_label.grid(row=3, column=1, sticky=tkinter.N + tkinter.W)
        self.measuremet_label.grid(row=4, column=1, pady=20, sticky=tkinter.N + tkinter.W)
        self.equipment_list_label.grid(row=5, column=1, pady=20, sticky=tkinter.N)
        self.equipment_tree.grid(row=6, column=1, sticky=tkinter.N + tkinter.W)
        self.receiver_list_label.grid(row=7, column=1, pady=20, sticky=tkinter.N)
        self.receiver_tree.grid(row=8, column=1, sticky=tkinter.N + tkinter.W)
        self.barrier_tree.grid(row=9, column=1, sticky=tkinter.N + tkinter.W)

    def generateEqmtTree(self):
        try: # delete tree if already exists
            self.equipment_tree.delete(*self.equipment_tree.get_children())
            self.equipment_tree_rows = []
            for i in self.parent.func_vars.equipment_list:
                self.equipment_tree_rows.append([i.count, i.eqmt_tag, i.path, i.make, i.model, i.sound_level, i.sound_ref_dist, i.x_coord, i.y_coord, i.z_coord])

            for i, value in enumerate(self.equipment_tree_rows):
                self.equipment_tree.insert("", "end", values=value)

        except:
            self.equipment_tree_columns = ["count", "tag", "path", "make", "model", "sound_level", "sound_ref_dist", "x", "y", "z"]
            self.equipment_tree_rows = []
            self.maxWidths = []

            # create widths
            for item in self.equipment_tree_columns:
                self.maxWidths.append(self.myFont.measure(str(item)))

            #create wors with eqmt data
            for i in self.parent.func_vars.equipment_list:
                self.equipment_tree_rows.append([i.count, i.eqmt_tag, i.path, i.make, i.model, i.sound_level, i.sound_ref_dist, i.x_coord, i.y_coord, i.z_coord])
            
            #getting max widths
            for col_idx in range(len(self.equipment_tree_rows[0])):
                maxWidth = self.maxWidths[col_idx]
                for row in self.equipment_tree_rows:
                    currentWidth = self.myFont.measure(str(row[col_idx]))
                    if currentWidth > maxWidth:
                        maxWidth = currentWidth
                self.maxWidths[col_idx] = maxWidth

            #initialize tree
            self.equipment_tree = tkinter.ttk.Treeview(self, columns=self.equipment_tree_columns, show='headings')

            # add rows and colmns to tree
            for col in self.equipment_tree_columns:
                new_length = self.myFont.measure(str(col))
                self.equipment_tree.heading(col, text=col)
                self.equipment_tree.column(col, minwidth=50, width=new_length+50, stretch=0)       
            for i, value in enumerate(self.equipment_tree_rows):
                self.equipment_tree.insert("", "end", values=value)
                #sizing
                if i == len(self.equipment_tree_rows)-1:
                    for col in self.equipment_tree_columns:
                        if col in ("eqmt_tag", "model"):
                            width_mult = 10
                            self.equipment_tree.column(col, minwidth=20, width=len(value)*width_mult, stretch=0)          

    def generateRcvrTree(self):
        try: # delete tree if already exists
            self.receiver_tree.delete(*self.receiver_tree.get_children())
            self.receiver_tree_rows = []
            for i in self.parent.func_vars.receiver_list:
                self.receiver_tree_rows.append([i.r_name, i.x_coord, i.y_coord, i.z_coord, i.sound_limit, i.predicted_sound_level])
            for i, value in enumerate(self.receiver_tree_rows):
                self.receiver_tree.insert("", "end", values=value)

        except:
            self.receiver_tree_columns = ["R#", "x", "y", "z", "sound limit", "est. level"]
            self.receiver_tree_rows = []
            self.maxWidths = []

            #create widths
            for item in self.receiver_tree_columns:
                self.maxWidths.append(self.myFont.measure(str(item)))
            
            #create rows with rcvr data
            for i in self.parent.func_vars.receiver_list:
                self.receiver_tree_rows.append([i.r_name, i.x_coord, i.y_coord, i.z_coord, i.sound_limit, i.predicted_sound_level])
            print(self.receiver_tree_rows)

            #getting max widths
            for col_idx in range(len(self.receiver_tree_rows[0])):
                maxWidth = self.maxWidths[col_idx]
                for row in self.receiver_tree_rows:
                    currentWidth = self.myFont.measure(str(row[col_idx]))
                    if currentWidth > maxWidth:
                        maxWidth = currentWidth
                self.maxWidths[col_idx] = maxWidth

            # initializing receiver tree
            self.receiver_tree = tkinter.ttk.Treeview(self, columns=self.receiver_tree_columns, show='headings')
            
            # adding columns and rows
            for col, maxWidth in zip(self.receiver_tree_columns, self.maxWidths):
                self.receiver_tree.heading(col, text=col)
                self.receiver_tree.column(col, minwidth=50, width=maxWidth+20, stretch=0)       
            for i, value in enumerate(self.receiver_tree_rows):
                self.receiver_tree.insert("", "end", values=value)

        self.equipment_tree.bind('<ButtonRelease-1>', self.select_item_from_eqmt_tree)
        self.receiver_tree.bind('<ButtonRelease-1>', self.select_item_from_rcvr_tree)

    def generateBarrierTree(self):
        try: # delete tree if already exists
            self.barrier_tree.delete(*self.barrier_tree.get_children())
            self.barrier_tree_rows = []
            for i in self.parent.func_vars.barrier_list:
                self.barrier_tree_rows.append([i.barrier_name, i.x0_coord, i.y0_coord, i.z0_coord, i.x1_coord, i.y1_coord, i.z1_coord])          
            for i, value in enumerate(self.barrier_tree_rows):
                self.barrier_tree.insert("", "end", values=value, tags=self.myFont)

        except:
            self.barrier_tree_columns = ["barrier_name", "x0", "y0", "z0", "x1", "y1", "z1"]
            self.barrier_tree_rows = []
            self.maxWidths = []

            #create widths
            for item in self.barrier_tree_columns:
                self.maxWidths.append(self.myFont.measure(str(item)))

            #create rows with barrier data
            for i in self.parent.func_vars.barrier_list:
                self.barrier_tree_rows.append([i.barrier_name, i.x0_coord, i.y0_coord, i.z0_coord, i.x1_coord, i.y1_coord, i.z1_coord])
            
            #getting max widths
            for col_idx in range(len(self.barrier_tree_rows[0])):
                maxWidth = self.maxWidths[col_idx]
                for row in self.barrier_tree_rows:
                    currentWidth = self.myFont.measure(str(row[col_idx]))
                    if currentWidth > maxWidth:
                        maxWidth = currentWidth
                self.maxWidths[col_idx] = maxWidth

            # initializing barrier tree 
            self.barrier_tree = tkinter.ttk.Treeview(self, columns=self.barrier_tree_columns, show='headings')
            
            # adding columns and rows
            for col in self.barrier_tree_columns:
                new_length = self.myFont.measure(str(col))
                self.barrier_tree.heading(col, text=col)
                self.barrier_tree.column(col, minwidth=50, width=new_length+50, stretch=0)          
            for i, value in enumerate(self.barrier_tree_rows):
                self.barrier_tree.insert("", "end", values=value, tags=self.myFont)


        self.equipment_tree.bind('<ButtonRelease-1>', self.select_item_from_eqmt_tree)
        self.barrier_tree.bind('<ButtonRelease-1>', self.select_item_from_rcvr_tree)

    def update_est_noise_levels(self):
        for rcvr in self.parent.func_vars.receiver_list:
            sound_pressure = 0
            for eqmt in self.parent.func_vars.equipment_list:
                if eqmt.sound_ref_dist == 0:
                    sound_power = eqmt.sound_level
                else:
                    q = 2 #need to update this
                    r = eqmt.sound_ref_dist*0.308
                    lp = eqmt.sound_level
                    b = q/(4*math.pi*r**2)
                    sound_power = lp + abs(10*math.log10(b))
                distance = math.sqrt((rcvr.x_coord-eqmt.x_coord)**2 + (rcvr.y_coord - eqmt.y_coord)**2 + (rcvr.z_coord - eqmt.z_coord)**2)
                print("rcvr", rcvr.x_coord, rcvr.y_coord, rcvr.z_coord)
                print("eqmt", eqmt.x_coord, eqmt.y_coord, eqmt.z_coord)
                print(sound_power, distance)
                try:
                    spl = sound_power-20*math.log10(distance/3.28)-8
                except ValueError:
                    print('MATH DOMAIN ERROR OCCURED')
                    spl = 1000

                sound_pressure += 10**(spl/10)
            rcvr.predicted_sound_level = round(10*math.log10(sound_pressure),1)
            if rcvr.r_name == "R1":
                print(f"predicted sound level: {rcvr.predicted_sound_level}")
                print(f"distance: {distance}")

    def select_item_from_eqmt_tree(self, event):
        self.current_equipment = self.equipment_tree.focus()
        self.current_euqipment = self.equipment_tree.item(self.current_equipment)['values']
        print(self.current_euqipment)

    def select_item_from_rcvr_tree(self, event):
        self.current_receiver = self.receiver_tree.focus()
        self.current_receiver = self.receiver_tree.item(self.current_receiver)['values']
        print(self.current_receiver)

    def onExportListButton(self):
        wb = openpyxl.load_workbook(XL_FILEPATH, keep_vba=True, data_only=False)
        ws = wb['Input LwA_XYZ']

        for obj in self.parent.func_vars.equipment_list:
            for row in ws.iter_rows(max_row=100):
                print("row 1", row[1].value)
                if row[1].value == None:
                    print("row 1", row[1].value)
                    break
                if row[1].value.replace(" ","-") == obj.eqmt_tag.replace(" ", "-"):
                    print("rplacing eqmt")
                    print("row 9", row[9].value)
                    row[9].value = obj.x_coord
                    row[10].value = obj.y_coord

        for obj in self.parent.func_vars.receiver_list:
            for row in ws.iter_rows(max_row=100):
                if row[15].value == None:
                    break
                if row[15].value.replace(" ","-") == obj.r_name.replace(" ", "-"):
                    print("row 16", row[16].value)
                    row[16].value = obj.x_coord
                    row[17].value = obj.y_coord

        # saving scale
        ws['U20'] = self.parent.func_vars.known_distance_ft
        ws['V20'] = self.parent.func_vars.scale_line_distance_px

        print("saving")
        wb.save(filename=XL_FILEPATH)
        print("saved")
        # wb.close()

    def e1_unfocus(self, event):
        self.status_label.focus()

    def e1_select_all(self, event):
        e1_text = self.e1.get()
        self.e1.selection_range(0, len(e1_text))

class Main_Application(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self) # , parent
        self.parent = parent

        self.func_vars = FuncVars(self)
        self.editor = Editor(self)
        self.pane_toolbox = Pane_Toolbox(self)
        self.pane_eqmt_info = Pane_Eqmt_Info(self)

        self.editor.grid(row=0, rowspan=2, column=0, stick=tkinter.N)
        self.pane_toolbox.grid(row=0, column=1, padx=20, pady=20, stick=tkinter.N+tkinter.W)
        self.pane_eqmt_info.grid(row=1, column=1, padx=20, pady=20, stick=tkinter.N)

def main():
    root = tkinter.Tk()
    mainApp = Main_Application(root)
    mainApp.pack(side="top", fill="both", expand=True)
    root.geometry('+0+0') #puts window in top left
    root.mainloop()

if __name__ == "__main__":
    main()
