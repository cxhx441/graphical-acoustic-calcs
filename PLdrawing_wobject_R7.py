import tkinter
from tkinter import ttk
from PIL import ImageTk, Image
import openpyxl
import math
import random
import sys
#sys.path.append('C:/Users/craig/Dropbox/00 - Cloud Documents/06 - Code/01 - Python Scripts')
import utils
import numpy
import tkinter.font
import acoustics
import csv
import BarrierPlotExporter

BED_IMAGE_FILEPATH = "bed_image.png"
TOP_IMAGE_FILEPATH = "top_image.png"
XL_FILEPATH = 'Aegis San Rafael - PL - 2020.08.17.xlsm'
XL_FILEPATH_SAVE = XL_FILEPATH[0:-5] + ' - exported.xlsm'
DRAWING_FONT = ('Helvetica 15 bold')

TAKE_ARI_BARRIER = True
TAKE_OB_FRESNAL_BARRIER = False

# if TAKE_ARI_BARRIER == True and TAKE_OB_BARRIER == True:
#     print("ERROR CANNOT HAVE BOTH BARRIER CALCULATION METHODS SET TO TRUE")

#setting columns
#EQMT
        #open workbook
wb = openpyxl.load_workbook(XL_FILEPATH, data_only=True)
ws = wb['Input LwA_XYZ']
EQMT_COUNT = ws['A']
EQMT_TAG = ws['B']
PATH = ws['C']
MAKE = ws['D']
MODEL = ws['E']
HZ63 = ws['F']
HZ125 = ws['G']
HZ250 = ws['H']
HZ500 = ws['I']
HZ1000 = ws['J']
HZ2000 = ws['K']
HZ4000 = ws['L']
HZ8000 = ws['M']
SOUND_LEVEL = ws['N']
SOUND_REF_DIST = ws['O']
TESTED_Q = ws['P']
INSTALLED_Q = ws['R']
EQMT_INSERTION_LOSS = ws['S']
EQMT_X_COORD = ws['T']
EQMT_Y_COORD = ws['U']
EQMT_Z_COORD = ws['V']

#RCVRS
R_NAME = ws['Z']
REC_X_COORD = ws['AA']
REC_Y_COORD = ws['AB']
REC_Z_COORD = ws['AC']
SOUND_LIMIT = ws['AD']

#BARRIERS
BARRIER_NAME = ws['Z']
BAR_X0_COORD = ws['AA']
BAR_Y0_COORD = ws['AB']
BAR_Z0_COORD = ws['AC']
BAR_X1_COORD = ws['AD']
BAR_Y1_COORD = ws['AE']
BAR_Z1_COORD = ws['AF']

#SCALING
KNOWN_DISTANCE_FT = ws['AE20']
SCALE_LINE_DISTANCE_PX = ws['AF20']

class FuncVars(object):
    def __init__(self, parent):
        self.parent = parent

        #initialize eqmt list
        self.equipment_list = list()
        for count, eqmt_tag, path, make, model, sound_level, sound_ref_dist, tested_q, installed_q, insertion_loss, x_coord, y_coord, z_coord, hz63, hz125, hz250, hz500, hz1000, hz2000, hz4000, hz8000 in zip(EQMT_COUNT, EQMT_TAG, PATH, MAKE, MODEL, SOUND_LEVEL, SOUND_REF_DIST, TESTED_Q, INSTALLED_Q, EQMT_INSERTION_LOSS, EQMT_X_COORD, EQMT_Y_COORD, EQMT_Z_COORD, HZ63, HZ125, HZ250, HZ500, HZ1000, HZ2000, HZ4000, HZ8000):
            if count.value == "Number of Units": continue
            if count.value == None: break
            print(z_coord.value)
            self.equipment_list.append(Equipment(count.value, str(eqmt_tag.value), path.value, make.value, model.value, sound_level.value, sound_ref_dist.value, tested_q.value, installed_q.value, insertion_loss.value, x_coord.value, y_coord.value, z_coord.value, hz63.value, hz125.value, hz250.value, hz500.value, hz1000.value, hz2000.value, hz4000.value, hz8000.value))

        #initialize rcvr list
        self.receiver_list = list()
        for r_name, x_coord, y_coord, z_coord, sound_limit in zip(R_NAME, REC_X_COORD, REC_Y_COORD, REC_Z_COORD, SOUND_LIMIT):
            if r_name.value == "R#": continue
            if r_name.value == None: break
            self.receiver_list.append(Receiver(str(r_name.value), x_coord.value, y_coord.value, z_coord.value, sound_limit.value, "NA"))

        #initialize barrier list
        self.barrier_list = list()
        for barrier_name, x0_coord, y0_coord, z0_coord, x1_coord, y1_coord, z1_coord in zip(BARRIER_NAME, BAR_X0_COORD, BAR_Y0_COORD, BAR_Z0_COORD, BAR_X1_COORD, BAR_Y1_COORD, BAR_Z1_COORD):
            if int(barrier_name.coordinate[1:]) < 24: continue
            if barrier_name.value == None: break
            self.barrier_list.append(Barrier(str(barrier_name.value), x0_coord.value, y0_coord.value, z0_coord.value, x1_coord.value, y1_coord.value, z1_coord.value))

        #initialize ignore matrix
        column = 91 # 1-index based
        row = 2
        self.ignore_matrix = list()
        for eqmt_coord in range(len(self.equipment_list)):
            ignore_rcvrs_list = list()
            for rcvr_coord in range(len(self.receiver_list)):
                ignore_rcvrs_list.append(ws.cell(row=row+eqmt_coord, column=column+rcvr_coord).value)
            self.ignore_matrix.append(ignore_rcvrs_list)
        # print("ignorematrix")
        # print(self.ignore_matrix)
        # print(len(self.ignore_matrix))

        #initialize master_scale
        self.old_master_scale = 1.0
        self.known_distance_ft = KNOWN_DISTANCE_FT.value if KNOWN_DISTANCE_FT.value != None else 1.0
        self.scale_line_distance_px = SCALE_LINE_DISTANCE_PX.value if SCALE_LINE_DISTANCE_PX.value != None else 1.0
        self.master_scale = self.known_distance_ft / self.scale_line_distance_px

    def update_master_scale(self, scale_line_distance_px, known_distance_ft):
        self.scale_line_distance_px = scale_line_distance_px
        self.known_distance_ft = known_distance_ft
        self.old_master_scale = self.master_scale
        self.master_scale = self.known_distance_ft / self.scale_line_distance_px

        '''rescaling eqmt'''
        for obj in self.equipment_list:
            obj.x_coord /= self.old_master_scale
            obj.y_coord /= self.old_master_scale
            obj.x_coord *= self.master_scale
            obj.y_coord *= self.master_scale
            obj.x_coord = round(obj.x_coord, 2)
            obj.y_coord = round(obj.y_coord, 2)

        for obj in self.receiver_list:
            obj.x_coord /= self.old_master_scale
            obj.y_coord /= self.old_master_scale
            obj.x_coord *= self.master_scale
            obj.y_coord *= self.master_scale
            obj.x_coord = round(obj.x_coord, 2)
            obj.y_coord = round(obj.y_coord, 2)
        '''rescaling eqmt'''

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateRcvrTree()
        self.parent.pane_eqmt_info.generateEqmtTree()

class Equipment(object):
    def __init__(self, count, eqmt_tag, path, make, model, sound_level, sound_ref_dist, tested_q, installed_q, insertion_loss, x_coord, y_coord, z_coord, hz63, hz125, hz250, hz500, hz1000, hz2000, hz4000, hz8000):
        self.count = count
        self.eqmt_tag = eqmt_tag.replace(" ", "-")
        self.path = path
        self.make = make
        self.model = model
        self.sound_level = sound_level if sound_level != None else 0
        self.sound_ref_dist = sound_ref_dist if sound_ref_dist != None else 0
        self.tested_q = tested_q
        self.installed_q = installed_q
        self.insertion_loss = insertion_loss if insertion_loss != None else 0
        self.x_coord = x_coord if x_coord != None else 0
        self.y_coord = y_coord if y_coord != None else 0
        self.z_coord = z_coord if z_coord != None else 0
        self.hz63 = hz63
        self.hz125 = hz125
        self.hz250 = hz250
        self.hz500 = hz500
        self.hz1000 = hz1000
        self.hz2000 = hz2000
        self.hz4000 = hz4000
        self.hz8000 = hz8000

class Receiver(object):
    def __init__(self, r_name, x_coord, y_coord, z_coord, sound_limit, predicted_sound_level):
        self.r_name = r_name.replace(" ", "-")
        self.x_coord = x_coord if x_coord != None else 0
        self.y_coord = y_coord if y_coord != None else 0
        self.z_coord = z_coord if z_coord != None else 0
        self.sound_limit = sound_limit
        self.predicted_sound_level = predicted_sound_level

class Barrier(object):
    def __init__(self, barrier_name, x0_coord, y0_coord, z0_coord, x1_coord, y1_coord, z1_coord):
        self.barrier_name = barrier_name.replace(" ", "-")
        self.x0_coord = x0_coord if x0_coord != None else 0
        self.y0_coord = y0_coord if y0_coord != None else 0
        self.z0_coord = z0_coord if z0_coord != None else 0
        self.x1_coord = x1_coord if x1_coord != None else 0
        self.y1_coord = y1_coord if y1_coord != None else 0
        self.z1_coord = z1_coord if z1_coord != None else 0

class Editor(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, parent)
        self.parent = parent

        #open image
        self.image = Image.open(BED_IMAGE_FILEPATH)
        self.image2 = Image.open(TOP_IMAGE_FILEPATH)

        #image sizing
        self.imageWidth, self.imageHeight = self.image.size
        print(self.image.size)
        self.image_size_factor = 1.5
        self.imageWidth *= self.image_size_factor
        self.imageHeight *= self.image_size_factor
        self.imageWidth = int(self.imageWidth)
        self.imageHeight = int(self.imageHeight)
        self.image = self.image.resize((self.imageWidth, self.imageHeight), Image.LANCZOS)
        self.image2_new_width = int(self.image2.size[0]/2)
        self.image2_new_height = int(self.image2.size[1]/2)
        self.image2 = self.image2.resize((self.image2_new_width, self.image2_new_height), Image.LANCZOS)
        self.tk_image = ImageTk.PhotoImage(self.image)
        self.tk_image2 = ImageTk.PhotoImage(self.image2)
        #self.tk_image2 = ImageTk.PhotoImage(self.image2.rotate(100))

        #canvas sizing
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.canvas_size_factor = 1
        self.canvasWidth = self.screen_width * self.canvas_size_factor
        self.canvasHeight = self.screen_height * self.canvas_size_factor
        self.canvasWidth -= 2000 # otherwise window is off the screen on home pc
        self.canvasHeight -= 250 # otherwise window is off the screen on home pc
        self.canvas = tkinter.Canvas(self, width=self.canvasWidth , height=self.canvasHeight, cursor="cross")

        self.canvas.config(scrollregion=(0, 0, self.imageWidth, self.imageHeight)) #giving scrollbars
        self.canvas.create_image(0,0, anchor="nw", image=self.tk_image, tag="bed_layer")
        image2_x_coord = self.image2.size[0]/2
        image2_y_coord = self.image2.size[1]/2
        self.canvas.create_image(image2_x_coord, image2_y_coord, tag="eqmt_drawing", image=self.tk_image2)

        '''scroll bar setup'''
        self.vScrollbar = tkinter.Scrollbar(self, orient=tkinter.VERTICAL)
        self.hScrollbar = tkinter.Scrollbar(self, orient=tkinter.HORIZONTAL)
        self.vScrollbar.config(command=self.canvas.yview)
        self.hScrollbar.config(command=self.canvas.xview)
        self.canvas.config(yscrollcommand=self.vScrollbar.set)
        self.canvas.config(xscrollcommand=self.hScrollbar.set)

        self.canvas.grid(row=0, column=0, sticky=tkinter.N + tkinter.S + tkinter.E + tkinter.W)
        self.vScrollbar.grid(row=0, column=1, stick=tkinter.N + tkinter.S)
        self.hScrollbar.grid(row=1, column=0, sticky=tkinter.E + tkinter.W)
        '''scroll bar setup'''

        self.initialize_eqmt_rcvr_barrier_drawings()

        self.temp_rect = None
        self.temp_line = None
        self.scale_line = None
        self.measure_line = None
        self.angle = 0

        self.canvas.bind("<Shift-ButtonPress-1>", self.shift_click)
        self.canvas.bind("<Shift-B1-Motion>", self.shift_click_move)
        self.canvas.bind("<Shift-ButtonRelease-1>", self.shift_click_release)

        '''Scrollable image'''
        self.canvas.bind('<Enter>', self._bound_to_mousewheel)
        self.canvas.bind('<Leave>', self._unbound_to_mousewheel)

    def initialize_eqmt_rcvr_barrier_drawings(self):
        '''initialize receivers and equipment boxes and barriers'''
        fontsize = 10
        for eqmt in self.parent.func_vars.equipment_list:
            # random_8bit_color = CraigsFunFunctions.random_8bit_color()
            green_hex_color = utils.rgb_to_hex((0, 254, 0))
            offset = 20
            x = eqmt.x_coord/self.parent.func_vars.master_scale
            y = eqmt.y_coord/self.parent.func_vars.master_scale
            # self.canvas.coords(self.temp_rect, self.x0-10, self.y0-10, self.curX+10, self.curY+10)
            self.rectPerm = self.canvas.create_rectangle(x-offset, y-offset, x+offset, y+offset, tag=eqmt.eqmt_tag, fill=green_hex_color, activeoutline='red')
            self.canvas.create_text(x, y, tag=eqmt.eqmt_tag, text=eqmt.eqmt_tag, font=DRAWING_FONT, fill='Black')

        for rcvr in self.parent.func_vars.receiver_list:
            # random_8bit_color = CraigsFunFunctions.random_8bit_color()
            red_hex_color = utils.rgb_to_hex((254, 0, 0))
            offset = 20
            x = rcvr.x_coord/self.parent.func_vars.master_scale
            y = rcvr.y_coord/self.parent.func_vars.master_scale
            # self.canvas.coords(self.temp_rect, self.x0-10, self.y0-10, self.curX+10, self.curY+10)
            self.rectPerm = self.canvas.create_rectangle(x-offset, y-offset, x+offset, y+offset, tag=rcvr.r_name, fill=red_hex_color, activeoutline='red')
            self.canvas.create_text(x, y, tag=rcvr.r_name, text=rcvr.r_name, font=DRAWING_FONT, fill='Black')

        for bar in self.parent.func_vars.barrier_list:
            x0 = bar.x0_coord/self.parent.func_vars.master_scale
            y0 = bar.y0_coord/self.parent.func_vars.master_scale
            x1 = bar.x1_coord/self.parent.func_vars.master_scale
            y1 = bar.y1_coord/self.parent.func_vars.master_scale
            self.linePerm = self.canvas.create_line(x0, y0, x1, y1, tag=bar.barrier_name, fill="purple", width=5)
            self.canvas.create_text(x0 + (x1-x0)/2, y0 + (y1 - y0)/2, tag=bar.barrier_name, text=bar.barrier_name, font=DRAWING_FONT, fill='Black')
        '''initialize receivers and equipment boxes and barriers'''

    def _bound_to_mousewheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Shift-MouseWheel>", self._on_shift_mousewheel)
    def _unbound_to_mousewheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")
    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    def _on_shift_mousewheel(self, event):
        self.canvas.xview_scroll(int(-1*(event.delta/120)), "units")
        '''Scrollable image'''

    def get_angle(self, x, y):
        v0 = [x, 0]
        v1 = [x, y]
        dot_product = numpy.dot(v0, v1)
        v0_mag = numpy.linalg.norm(v0)
        v1_mag = numpy.linalg.norm(v1)
        angle = math.degrees(numpy.arccos((dot_product/(v0_mag*v1_mag))))
        if dot_product < 0:
            angle += 90
        print("hey", angle)
        return angle
    def update_distance_label(self):
        self.parent.pane_eqmt_info.measurement_label.configure(text="Measurement: " + str(round(self.parent.func_vars.master_scale*(math.sqrt((self.x0 - self.curX)**2 + (self.y0 - self.curY)**2)),2)) + " ft")
    def get_current_n_start_mouse_pos(self, event):
        self.x0 = self.canvas.canvasx(event.x)
        self.y0 = self.canvas.canvasy(event.y)
        self.curX = self.canvas.canvasx(event.x)
        self.curY = self.canvas.canvasy(event.y)
    def get_current_mouse_pos(self, event):
        self.curX = self.canvas.canvasx(event.x)
        self.curY = self.canvas.canvasy(event.y)

    def drawing_grid_leftMouseClick(self, event):
        self.canvas.delete("grid_rect")
        self.canvas.delete("grid_level")
        self.get_current_n_start_mouse_pos(event)
        self.temp_rect = self.canvas.create_rectangle(self.x0, self.y0, self.x0, self.y0, outline='green', width=5)
    def drawing_grid_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_rect, self.x0, self.y0, self.curX, self.curY)
    def drawing_grid_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_rect)
        self.grid_rect = self.canvas.create_rectangle(self.x0, self.y0, self.curX, self.curY, outline='green', width=5, tag="grid_rect")

    def setting_scale_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)

        if self.scale_line != None:
            self.canvas.delete(self.scale_line)
        self.temp_scale_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="orange", width=5)
    def setting_scale_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_scale_line, self.x0, self.y0, self.curX, self.curY)
    def setting_scale_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_scale_line)

        self.scale_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="blue", width=5)
        scale_line_coords = self.canvas.coords(self.scale_line)
        # self.parent.func_vars.scale_line_distance_px = CraigsFunFunctions.distance_formula(scale_line_coords[0], scale_line_coords[2], scale_line_coords[1], scale_line_coords[3])
        # self.parent.func_vars.known_distance_ft = float(self.parent.pane_eqmt_info.e1.get())
        _scale_line_distance_px = utils.distance_formula(scale_line_coords[0], scale_line_coords[2], scale_line_coords[1], scale_line_coords[3])
        _known_distance_ft = float(self.parent.pane_eqmt_info.e1.get())
        self.parent.func_vars.update_master_scale(_scale_line_distance_px, _known_distance_ft)
        # self.parent.func_vars.old_master_scale = self.parent.func_vars.master_scale
        # self.parent.func_vars.master_scale = self.parent.func_vars.known_distance_ft / self.parent.func_vars.scale_line_distance_px

        scaleIndicatorLabelText = "Scale: " + str(round(self.parent.func_vars.scale_line_distance_px,0)) + " px = " + str(self.parent.func_vars.known_distance_ft) + " ft"
        self.parent.pane_eqmt_info.scaleIndicatorLabel.configure(text=scaleIndicatorLabelText)

    def drawing_eqmt_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.temp_rect = self.canvas.create_rectangle(self.x0, self.y0, self.x0, self.y0, outline='red')
    def drawing_eqmt_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_rect, self.x0, self.y0, self.curX, self.curY)
    def drawing_eqmt_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_rect)

        # random_8bit_color = CraigsFunFunctions.random_8bit_color()
        green_hex_color = utils.rgb_to_hex((0, 254, 0))

        eqmt_tag=self.parent.pane_eqmt_info.current_equipment[1] # i think this is grabbing from the tree
        tagged_objects = self.canvas.find_withtag(eqmt_tag)
        for tagged_object in tagged_objects:
            self.canvas.delete(tagged_object)
        self.rectPerm = self.canvas.create_rectangle(self.x0, self.y0, self.curX, self.curY, tag=eqmt_tag, fill=green_hex_color, activeoutline='red')

        self.canvas.create_text((self.x0 + (self.curX-self.x0)/2, self.y0 + (self.curY - self.y0)/2), tag=eqmt_tag, text=eqmt_tag, font=DRAWING_FONT, fill='Black')

        #update this one piece of eqmt
        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == eqmt_tag:
                obj.x_coord = self.x0 + (self.curX - self.x0)/2
                obj.y_coord = self.y0 + (self.curY - self.y0)/2
                obj.x_coord *= self.parent.func_vars.master_scale
                obj.y_coord *= self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)
                # print(obj.x_coord)
                # print(obj.y_coord)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateRcvrTree()
        self.parent.pane_eqmt_info.generateEqmtTree()

    def drawing_rcvr_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.temp_rect = self.canvas.create_rectangle(self.x0, self.y0, self.x0, self.y0, outline='green')
    def drawing_rcvr_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_rect, self.x0, self.y0, self.curX, self.curY)
    def drawing_rcvr_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_rect)

        # random_8bit_color = CraigsFunFunctions.random_8bit_color()
        red_hex_color = utils.rgb_to_hex((254, 0, 0))

        r_name=self.parent.pane_eqmt_info.current_receiver[0]
        tagged_objects = self.canvas.find_withtag(r_name)
        for tagged_object in tagged_objects:
            self.canvas.delete(tagged_object)
        self.rectPerm = self.canvas.create_rectangle(self.x0, self.y0, self.curX, self.curY, tag=r_name, fill=red_hex_color, activeoutline='red')

        self.canvas.create_text((self.x0 + (self.curX-self.x0)/2, self.y0 + (self.curY - self.y0)/2), tag=r_name, text=r_name, font=DRAWING_FONT, fill='Black')

        #update this one rcvr
        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == r_name:
                obj.x_coord = self.x0 + (self.curX - self.x0)/2
                obj.y_coord = self.y0 + (self.curY - self.y0)/2
                obj.x_coord *= self.parent.func_vars.master_scale
                obj.y_coord *= self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateRcvrTree()

    def drawing_barrier_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.temp_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill='yellow', width=5)
    def drawing_barrier_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_line, self.x0, self.y0, self.curX, self.curY)
    def drawing_barrier_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_line)

        barrier_name=self.parent.pane_eqmt_info.current_barrier[0]
        tagged_objects = self.canvas.find_withtag(barrier_name)
        for tagged_object in tagged_objects:
            self.canvas.delete(tagged_object)
        self.barPerm = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, tag=barrier_name, fill="purple", width=5)

        self.canvas.create_text((self.x0 + (self.curX-self.x0)/2, self.y0 + (self.curY - self.y0)/2), tag=barrier_name, text=barrier_name, font=DRAWING_FONT, fill='Black')

        #update this one bar
        for obj in self.parent.func_vars.barrier_list:
            if obj.barrier_name == barrier_name:
                obj.x0_coord = self.x0
                obj.y0_coord = self.y0
                obj.x1_coord = self.curX
                obj.y1_coord = self.curY
                obj.x0_coord *= self.parent.func_vars.master_scale
                obj.y0_coord *= self.parent.func_vars.master_scale
                obj.x1_coord *= self.parent.func_vars.master_scale
                obj.y1_coord *= self.parent.func_vars.master_scale
                obj.x0_coord = round(obj.x0_coord, 2)
                obj.y0_coord = round(obj.y0_coord, 2)
                obj.x1_coord = round(obj.x1_coord, 2)
                obj.y1_coord = round(obj.y1_coord, 2)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateRcvrTree()
        self.parent.pane_eqmt_info.generateBarrierTree()

    def measureing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        if self.measure_line != None:
            self.canvas.delete(self.measure_line)
        self.update_distance_label()
        self.temp_measure_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="orange", width=5)
    def measureing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.coords(self.temp_measure_line, self.x0, self.y0, self.curX, self.curY)
        self.update_distance_label()
    def measureing_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.canvas.delete(self.temp_measure_line)
        self.measure_line = self.canvas.create_line(self.x0, self.y0, self.curX, self.curY, fill="red", width=5)

    def rotating_eqmt_drawing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        #calc angle at start point
        self.eqmt_drawing_center = self.canvas.coords("eqmt_drawing")
        self.angle0 = self.get_angle(self.x0-self.eqmt_drawing_center[0], self.y0-self.eqmt_drawing_center[1]) - self.angle
    def rotating_eqmt_drawing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        # calculate current angle relative to initial angle
        self.angle_1 = self.get_angle(self.curX-self.eqmt_drawing_center[0], self.curY-self.eqmt_drawing_center[1])
        dwg_x = self.eqmt_drawing_center[0]
        dwg_y = self.eqmt_drawing_center[1]

        if self.curX > dwg_x and self.curY > dwg_y:
            self.angle_1 *= -1
        elif self.curX < dwg_x and self.curY < dwg_y:
            self.angle_1 = 180 - self.angle_1
        elif self.curX < dwg_x and self.curY > dwg_y:
            self.angle_1 = self.angle_1 + 180
        elif self.curX < dwg_x and self.curY > dwg_y:
            self.angle_1 = self.angle_1 + 360

        self.angle = self.angle_1

        self.canvas.delete("eqmt_drawing")
        self.tk_image2 = ImageTk.PhotoImage(self.image2.rotate(self.angle, expand=True))
        self.canvas.create_image(self.eqmt_drawing_center[0], self.eqmt_drawing_center[1], image=self.tk_image2, tag="eqmt_drawing")
        self.canvas.tag_lower("eqmt_rdawing")
        self.canvas.tag_lower("bed_layer")

    def moving_eqmt_drawing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.eqmt_drawing_center = self.canvas.coords("eqmt_drawing")
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.delete("eqmt_drawing")
        self.canvas.create_image(self.eqmt_drawing_center[0] + x_shifter, self.eqmt_drawing_center[1] + y_shifter, image=self.tk_image2, tag="eqmt_drawing")
    def moving_eqmt_drawing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        # self.eqmt_drawing_center = self.canvas.coords("eqmt_drawing")
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.delete("eqmt_drawing")
        self.canvas.create_image(self.eqmt_drawing_center[0] + x_shifter, self.eqmt_drawing_center[1] + y_shifter, image=self.tk_image2, tag="eqmt_drawing")
        self.canvas.tag_lower("eqmt_drawing")
        self.canvas.tag_lower("bed_layer")

    def resizing_eqmt_drawing_leftMouseClick(self, event):
        self.get_current_n_start_mouse_pos(event)
        self.eqmt_dwg_cntr = self.canvas.coords("eqmt_drawing")
        self.eqmt_dwg_width_0 = self.image2_new_width
        self.eqmt_dwg_height_0 = self.image2_new_height
        self.eqmt_dwg_ratio = self.eqmt_dwg_width_0 / self.eqmt_dwg_height_0
        self.rect_p1_x0 = self.eqmt_dwg_cntr[0]-self.eqmt_dwg_width_0/2
        self.rect_p2_x0 = self.eqmt_dwg_cntr[0]+self.eqmt_dwg_width_0/2
        self.rect_p1_y0 = self.eqmt_dwg_cntr[1]-self.eqmt_dwg_height_0/2
        self.rect_p2_y0 = self.eqmt_dwg_cntr[1]+self.eqmt_dwg_height_0/2
        self.temp_rect = self.canvas.create_rectangle(self.rect_p1_x0, self.rect_p1_y0, self.rect_p2_x0, self.rect_p2_y0, outline='red')
    def resizing_eqmt_drawing_leftMouseMove(self, event):
        self.get_current_mouse_pos(event)
        self.x_change = self.curX - self.x0
        self.y_change = self.curY - self.y0
        self.rect_p1_x1 = self.rect_p1_x0
        self.rect_p2_x2 = self.rect_p2_x0 + self.x_change
        self.rect_p1_y1 = self.rect_p1_y0
        self.rect_p2_y1 = self.rect_p2_y0 + self.x_change / self.eqmt_dwg_ratio
        self.canvas.coords(self.temp_rect, self.rect_p1_x1, self.rect_p1_y1, self.rect_p2_x2, self.rect_p2_y1)
    def resizing_eqmt_drawing_leftMouseRelease(self, event):
        self.get_current_mouse_pos(event)
        self.eqmt_dwg_width_1 = self.eqmt_dwg_width_0 + int(self.x_change)
        self.eqmt_dwg_height_1 = int(self.eqmt_dwg_width_1 / self.eqmt_dwg_ratio)
        self.eqmt_dwg_cntr[0] += int(self.x_change/2)
        self.eqmt_dwg_cntr[1] += int((self.x_change / self.eqmt_dwg_ratio)/2)

        self.canvas.delete("eqmt_drawing")
        self.canvas.delete(self.temp_rect)
        self.image2 = self.image2.resize((self.eqmt_dwg_width_1, self.eqmt_dwg_height_1), Image.LANCZOS)
        self.tk_image2 = ImageTk.PhotoImage(self.image2.rotate(self.angle, expand=True))
        self.canvas.create_image(self.eqmt_dwg_cntr[0], self.eqmt_dwg_cntr[1], image=self.tk_image2, tag="eqmt_drawing")

        self.image2_new_width = self.eqmt_dwg_width_1
        self.image2_new_height = self.eqmt_dwg_height_1

        self.canvas.tag_lower("eqmt_drawing")
        self.canvas.tag_lower("bed_layer")

    def shift_click(self, event):
        if self.canvas.find_withtag("current"):
            self.eqmt_rcvr_or_barr_tagged = self.canvas.gettags("current")
            self.tag_rcvr_or_barr_num = self.eqmt_rcvr_or_barr_tagged[0]
            self.eqmt_rcvr_barr_ids = self.canvas.find_withtag(self.eqmt_rcvr_or_barr_tagged[0])
            self.current_shape = self.eqmt_rcvr_barr_ids[0]
            self.current_text = self.eqmt_rcvr_barr_ids[1]
            self.current_shape_coords = self.canvas.coords(self.current_shape)
            self.current_text_coords = self.canvas.coords(self.current_text)

            self.get_current_n_start_mouse_pos(event)

        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == self.tag_rcvr_or_barr_num:
                self.obj_x_coord_0 = obj.x_coord
                self.obj_y_coord_0 = obj.y_coord

        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == self.tag_rcvr_or_barr_num:
                self.obj_x_coord_0 = obj.x_coord
                self.obj_y_coord_0 = obj.y_coord

        for obj in self.parent.func_vars.barrier_list:
            if obj.barrier_name == self.tag_rcvr_or_barr_num:
                self.obj_x_coord_0 = obj.x0_coord
                self.obj_y_coord_0 = obj.y0_coord
                self.obj_x_coord_1 = obj.x1_coord
                self.obj_y_coord_1 = obj.y1_coord

    def shift_click_move(self, event):
        self.get_current_mouse_pos(event)
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.coords(self.current_shape, self.current_shape_coords[0]+x_shifter, self.current_shape_coords[1]+y_shifter, self.current_shape_coords[2]+x_shifter, self.current_shape_coords[3]+y_shifter)
        self.canvas.coords(self.current_text, self.current_text_coords[0]+x_shifter, self.current_text_coords[1]+y_shifter)

        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == self.tag_rcvr_or_barr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == self.tag_rcvr_or_barr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        for obj in self.parent.func_vars.barrier_list:
            if obj.barrier_name == self.tag_rcvr_or_barr_num:
                obj.x0_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y0_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x1_coord = self.obj_x_coord_1 + x_shifter*self.parent.func_vars.master_scale
                obj.y1_coord = self.obj_y_coord_1 + y_shifter*self.parent.func_vars.master_scale
                obj.x0_coord = round(obj.x0_coord, 2)
                obj.y0_coord = round(obj.y0_coord, 2)
                obj.x1_coord = round(obj.x1_coord, 2)
                obj.y1_coord = round(obj.y1_coord, 2)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateEqmtTree()
        self.parent.pane_eqmt_info.generateRcvrTree()
        self.parent.pane_eqmt_info.generateBarrierTree()
    def shift_click_release(self, event):
        self.get_current_mouse_pos(event)
        x_shifter = self.curX - self.x0
        y_shifter = self.curY - self.y0
        self.canvas.coords(self.current_shape, self.current_shape_coords[0]+x_shifter, self.current_shape_coords[1]+y_shifter, self.current_shape_coords[2]+x_shifter, self.current_shape_coords[3]+y_shifter)
        self.canvas.coords(self.current_text, self.current_text_coords[0]+x_shifter, self.current_text_coords[1]+y_shifter)

        for obj in self.parent.func_vars.equipment_list:
            if obj.eqmt_tag == self.tag_rcvr_or_barr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        for obj in self.parent.func_vars.receiver_list:
            if obj.r_name == self.tag_rcvr_or_barr_num:
                obj.x_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x_coord = round(obj.x_coord, 2)
                obj.y_coord = round(obj.y_coord, 2)

        for obj in self.parent.func_vars.barrier_list:
            if obj.barrier_name == self.tag_rcvr_or_barr_num:
                obj.x0_coord = self.obj_x_coord_0 + x_shifter*self.parent.func_vars.master_scale
                obj.y0_coord = self.obj_y_coord_0 + y_shifter*self.parent.func_vars.master_scale
                obj.x1_coord = self.obj_x_coord_1 + x_shifter*self.parent.func_vars.master_scale
                obj.y1_coord = self.obj_y_coord_1 + y_shifter*self.parent.func_vars.master_scale
                obj.x0_coord = round(obj.x0_coord, 2)
                obj.y0_coord = round(obj.y0_coord, 2)
                obj.x1_coord = round(obj.x1_coord, 2)
                obj.y1_coord = round(obj.y1_coord, 2)

        self.parent.pane_eqmt_info.update_est_noise_levels()
        self.parent.pane_eqmt_info.generateEqmtTree()
        self.parent.pane_eqmt_info.generateRcvrTree()
        self.parent.pane_eqmt_info.generateBarrierTree()

class Pane_Toolbox(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, parent)
        self.parent = parent

        self.button_set_image_scale = tkinter.Button(self, text="Set Image Scale", command=self.set_scale, font=(None, 15))
        self.button_measure = tkinter.Button(self, text="Measure", command=self.measure, font=(None, 15))
        self.button_draw_equipment = tkinter.Button(self, text="Draw Equipment", command=self.draw_equipment, font=(None, 15))
        self.button_draw_receiver = tkinter.Button(self, text="Draw Receiver", command=self.draw_receiver, font=(None, 15))
        self.button_draw_barrier = tkinter.Button(self, text="Draw Barrier", command=self.draw_barrier, font=(None, 15))
        self.button_rotate_eqmt_drawing = tkinter.Button(self, text="Eqmt Drawing - Rotate", command=self.rotate_eqmt_drawing, font=(None, 15))
        self.button_move_eqmt_drawing = tkinter.Button(self, text="Eqmt Drawing - Move", command=self.move_eqmt_drawing, font=(None, 15))
        self.button_resize_eqmt_drawing = tkinter.Button(self, text="Eqmt Drawing - Resize", command=self.resize_eqmt_drawing, font=(None, 15))
        self.button_draw_grid = tkinter.Button(self, text="Draw Grid", command=self.draw_grid, font=(None, 15))
        self.button_update_grid = tkinter.Button(self, text="Update Grid", command=self.update_grid, font=(None, 15))
        self.button_export_bar_file = tkinter.Button(self, text="Export Bar to File", command=self.export_bar_file, font=(None, 15))

        self.button_set_image_scale.grid(row=0, column=0, sticky=tkinter.N + tkinter.W)
        self.button_measure.grid(row=1, column=0, sticky=tkinter.N + tkinter.W)
        self.button_draw_equipment.grid(row=0, column=1, sticky=tkinter.N + tkinter.W)
        self.button_draw_receiver.grid(row=1, column=1, sticky=tkinter.N + tkinter.W)
        self.button_draw_barrier.grid(row=2, column=1, sticky=tkinter.N + tkinter.W)
        self.button_rotate_eqmt_drawing.grid(row=0, column=2, sticky=tkinter.N + tkinter.W)
        self.button_move_eqmt_drawing.grid(row=1, column=2, sticky=tkinter.N + tkinter.W)
        self.button_resize_eqmt_drawing.grid(row=2, column=2, sticky=tkinter.N + tkinter.W)
        self.button_draw_grid.grid(row=0, column=3, sticky=tkinter.N + tkinter.W)
        self.button_update_grid.grid(row=1, column=3, sticky=tkinter.N + tkinter.W)
        self.button_export_bar_file.grid(row=2, column=3, sticky=tkinter.N + tkinter.W)

    def export_bar_file(self):
        with open('bar_export_list.csv', mode='w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            for barrier_item in self.parent.pane_eqmt_info.barrierListForExcelOutput:
                print(barrier_item)
                csv_writer.writerow(barrier_item)
        BarrierPlotExporter.exportBarrierPlots(self.parent.pane_eqmt_info.barrierListForExcelOutput[1:])

    def draw_grid(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.drawing_grid_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.drawing_grid_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.drawing_grid_leftMouseRelease)

        self.parent.pane_eqmt_info.status_label.configure(text='Status: Drawing Grid')
        self.parent.pane_eqmt_info.e1.delete(0, 'end')
        self.parent.pane_eqmt_info.e1.insert(0, "elevation, spacing (ft)")

        self.parent.pane_eqmt_info.e1.focus()

    def update_grid(self):
        self.parent.editor.canvas.delete("grid_level")
        inputdata = self.parent.pane_eqmt_info.e1.get()
        inputdata_list = inputdata.split(", ")
        grid_elevation = int(inputdata_list[0])
        spacing = int(inputdata_list[1])
        grid_rect_coords = self.parent.editor.canvas.coords(self.parent.editor.grid_rect)
        start_x_coord_ft = grid_rect_coords[0]*self.parent.func_vars.master_scale
        start_y_coord_ft = grid_rect_coords[1]*self.parent.func_vars.master_scale
        end_x_coord_ft = grid_rect_coords[2]*self.parent.func_vars.master_scale
        end_y_coord_ft = grid_rect_coords[3]*self.parent.func_vars.master_scale

        grid_receiver_list = []
        cur_x_coord_ft = start_x_coord_ft
        cur_y_coord_ft = start_y_coord_ft
        while cur_y_coord_ft < end_y_coord_ft:
            while cur_x_coord_ft < end_x_coord_ft:
                grid_receiver_list.append([cur_x_coord_ft, cur_y_coord_ft, "0"])
                cur_x_coord_ft += spacing
            cur_y_coord_ft += spacing
            cur_x_coord_ft = start_x_coord_ft
        print(grid_receiver_list)

        #calculating noise levels at receiver in grid list
        for grd_rcvr in grid_receiver_list:
            rcvr_x_coord = grd_rcvr[0]
            rcvr_y_coord = grd_rcvr[1]
            sound_pressure = 0
            for eqmt in self.parent.func_vars.equipment_list:
                if eqmt.sound_ref_dist == 0:
                    sound_power = eqmt.sound_level
                else:
                    q = eqmt.tested_q #need to update this
                    r = eqmt.sound_ref_dist*0.308
                    lp = eqmt.sound_level
                    b = q/(4*math.pi*r**2)
                    sound_power = lp + abs(10*math.log10(b))
                distance = math.sqrt((rcvr_x_coord-eqmt.x_coord)**2 + (rcvr_y_coord - eqmt.y_coord)**2 + (grid_elevation - eqmt.z_coord)**2)
                try:
                    q = eqmt.installed_q
                    r = distance*0.308
                    attenuation = abs(10*math.log10(q/(4*math.pi*r**2)))
                    used_barrier_name = None
                    barrier_IL = 0
                    if TAKE_ARI_BARRIER == True and TAKE_OB_FRESNAL_BARRIER == False:
                        for bar in self.parent.func_vars.barrier_list:
                            barrier_info_list = self.parent.pane_eqmt_info.ARI_barrier_IL_calc(eqmt.x_coord, eqmt.y_coord, eqmt.z_coord, bar.x0_coord, bar.y0_coord, bar.z0_coord, bar.x1_coord, bar.y1_coord, bar.z1_coord, rcvr_x_coord, rcvr_y_coord, grid_elevation)
                            barrier_IL_test = barrier_info_list[0] if barrier_info_list != 0 else 0
                            if barrier_IL_test > barrier_IL:
                                barrier_IL = barrier_IL_test
                                used_barrier_name = str(bar.barrier_name + ' - ari')

                    if TAKE_ARI_BARRIER == True and TAKE_OB_FRESNAL_BARRIER == True:
                        for bar in self.parent.func_vars.barrier_list:
                            if None not in [eqmt.hz63, eqmt.hz125, eqmt.hz250, eqmt.hz500, eqmt.hz1000, eqmt.hz2000, eqmt.hz4000, eqmt.hz8000]:
                                barrier_info_list = self.parent.pane_eqmt_info.OB_fresnel_barrier_IL_calc(eqmt.x_coord, eqmt.y_coord, eqmt.z_coord, eqmt.hz63, eqmt.hz125, eqmt.hz250, eqmt.hz500, eqmt.hz1000, eqmt.hz2000, eqmt.hz4000, eqmt.hz8000, eqmt.sound_level, bar.x0_coord, bar.y0_coord, bar.z0_coord, bar.x1_coord, bar.y1_coord, bar.z1_coord, rcvr_x_coord, rcvr_y_coord, grid_elevation)
                                barrier_IL_test = barrier_info_list[0] if barrier_info_list != 0 else 0
                                barriermethod = ' - OB_fresnel'
                            else:
                                barrier_info_list = self.parent.pane_eqmt_info.ARI_barrier_IL_calc(eqmt.x_coord, eqmt.y_coord, eqmt.z_coord, bar.x0_coord, bar.y0_coord, bar.z0_coord, bar.x1_coord, bar.y1_coord, bar.z1_coord, rcvr_x_coord, rcvr_y_coord, grid_elevation)
                                barrier_IL_test = barrier_info_list[0] if barrier_info_list != 0 else 0
                                barriermethod = ' - ari'
                            if barrier_IL_test > barrier_IL:

                                barrier_IL = barrier_IL_test
                                used_barrier_name = str(bar.barrier_name + barriermethod)

                    spl = sound_power-eqmt.insertion_loss-attenuation-barrier_IL
                except ValueError:
                    print('MATH DOMAIN ERROR OCCURED')
                    spl = 1000
                sound_pressure += 10**(spl/10)
            grd_rcvr[2] = str(round(10*math.log10(sound_pressure),1))

        colorscale = [x for x in range(35, 95, 10)]
        colorlist = ["black", "blue", "purple", "cyan3", "green3", "yellow3", "DarkOrange1", "OrangeRed2", "maroon2"]
        for grid_rcvr in grid_receiver_list:
            x = grid_rcvr[0]/self.parent.func_vars.master_scale
            y = grid_rcvr[1]/self.parent.func_vars.master_scale
            level = grid_rcvr[2]
            textcolor = "black"
            for colorrange, color in zip(colorscale, colorlist):
                consider_level = int(round(float(level), 0))
                print("colorrange", colorrange)
                print("considerlevel", consider_level)
                if consider_level >= colorrange:
                    textcolor = color

            self.parent.editor.canvas.create_text((x, y), tag="grid_level", text=str(consider_level), font=("arialbd.ttf", 10), fill=textcolor)

    def set_scale(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.setting_scale_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.setting_scale_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.setting_scale_leftMouseRelease)

        self.parent.pane_eqmt_info.status_label.configure(text='Status: Setting Scale')
        self.parent.pane_eqmt_info.e1.delete(0, 'end')
        self.parent.pane_eqmt_info.e1.insert(0, "distance (ft)")
        self.parent.pane_eqmt_info.e1.focus()

    def draw_equipment(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.drawing_eqmt_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.drawing_eqmt_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.drawing_eqmt_leftMouseRelease)

        self.parent.pane_eqmt_info.status_label.configure(text='Status: Drawing Equipment')

    def draw_receiver(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.drawing_rcvr_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.drawing_rcvr_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.drawing_rcvr_leftMouseRelease)
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Drawing Receiver')

    def draw_barrier(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.drawing_barrier_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.drawing_barrier_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.drawing_barrier_leftMouseRelease)
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Drawing Barrier')

    def measure(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.measureing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.measureing_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.measureing_leftMouseRelease)
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Measuring')

    def rotate_eqmt_drawing(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.rotating_eqmt_drawing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.rotating_eqmt_drawing_leftMouseMove)
        self.parent.editor.canvas.unbind("<ButtonRelease-1>")
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Rotating Equipment Drawing')

    def move_eqmt_drawing(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.moving_eqmt_drawing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.moving_eqmt_drawing_leftMouseMove)
        self.parent.editor.canvas.unbind("<ButtonRelease-1>")
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Moving Equipment Drawing')

    def resize_eqmt_drawing(self):
        self.parent.editor.canvas.bind("<ButtonPress-1>", self.parent.editor.resizing_eqmt_drawing_leftMouseClick)
        self.parent.editor.canvas.bind("<B1-Motion>", self.parent.editor.resizing_eqmt_drawing_leftMouseMove)
        self.parent.editor.canvas.bind("<ButtonRelease-1>", self.parent.editor.resizing_eqmt_drawing_leftMouseRelease)
        self.parent.pane_eqmt_info.status_label.configure(text='Status: Resizing Equipment Drawing')

class Pane_Eqmt_Info(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, parent)
        self.parent = parent
        self.update_est_noise_levels()

        self.myFont = tkinter.font.nametofont('TkTextFont')

        self.e1 = tkinter.Entry(self, font=(None, 15), width=36)
        self.e1.insert(0, "input scale & eqmt_tag names here prior to setting")
        self.e1.bind("<FocusIn>", self.e1_select_all)
        self.e1.bind("<Return>", self.e1_unfocus)

        scaleIndicatorLabelText = "Scale: " + str(round(self.parent.func_vars.scale_line_distance_px,0)) + " px = " + str(self.parent.func_vars.known_distance_ft) + " ft"

        self.exportList_button = tkinter.Button(self, text="Export Tag List", command=self.onExportListButton, font=(None, 15))
        self.scaleIndicatorLabel = tkinter.Label(self, text=scaleIndicatorLabelText, borderwidth=2, relief="solid", font=(None, 15))
        self.status_label = tkinter.Label(self, text="Status: Idle", borderwidth=2, relief="solid", font=(None, 15))
        self.measurement_label = tkinter.Label(self, text="Measurement: ", borderwidth=2, relief="solid", font=(None, 15))
        self.equipment_list_label = tkinter.Label(self, text="Equipment", font=(None, 15))
        self.receiver_list_label = tkinter.Label(self, text="Receivers", font=(None, 15))
        self.barrier_list_label = tkinter.Label(self, text="Barriers", font=(None, 15))
        self.generateEqmtTree()
        self.generateRcvrTree()
        self.generateBarrierTree()
        self.generateIgnoreMatrixTree()

        self.equipment_tree.bind("<Double-1>", self.open_item_editor_window)
        self.receiver_tree.bind("<Double-1>", self.open_item_editor_window)
        self.barrier_tree.bind("<Double-1>", self.open_item_editor_window)
        self.deselect_item_from_trees()

        self.e1.grid(row=0, column=0, sticky=tkinter.N + tkinter.W)
        self.exportList_button.grid(row=1, column=0, sticky=tkinter.N + tkinter.W)
        self.scaleIndicatorLabel.grid(row=2, column=0, sticky=tkinter.N + tkinter.W)
        self.status_label.grid(row=3, column=0, sticky=tkinter.N + tkinter.W)
        self.measurement_label.grid(row=4, column=0, sticky=tkinter.N + tkinter.W)
        self.equipment_list_label.grid(row=5, column=0, pady=10, sticky=tkinter.N + tkinter.W)
        self.equipment_tree.grid(row=6, column=0, sticky=tkinter.N + tkinter.W)
        self.receiver_list_label.grid(row=7, column=0, pady=10, sticky=tkinter.N+tkinter.W)
        self.receiver_tree.grid(row=8, column=0, sticky=tkinter.N + tkinter.W)
        self.barrier_list_label.grid(row=9, column=0, pady=10, sticky=tkinter.N+tkinter.W)
        self.barrier_tree.grid(row=10, column=0, sticky=tkinter.N + tkinter.W)
        self.ignore_matrix_tree.grid(row=8, column=0, padx=400, sticky=tkinter.N+tkinter.W)

    def generateEqmtTree(self):
        try: # delete tree if already exists
            self.equipment_tree.delete(*self.equipment_tree.get_children())
            self.equipment_tree_rows = []
            for i in self.parent.func_vars.equipment_list:
                self.equipment_tree_rows.append([i.count, i.eqmt_tag, i.path, i.make, i.model, i.sound_level, i.sound_ref_dist, i.tested_q, i.installed_q,  i.insertion_loss, i.x_coord, i.y_coord, i.z_coord])

            for i, value in enumerate(self.equipment_tree_rows):
                self.equipment_tree.insert("", "end", values=value, tags=self.myFont)

        except:
            self.equipment_tree_columns = ["count", "tag", "path", "make", "model", "sound_level", "sound_ref_dist", "Q (tested)", "Q (installed)", "IL", "x", "y", "z"]
            self.equipment_tree_rows = []
            self.maxWidths = []

            # create widths
            for item in self.equipment_tree_columns:
                self.maxWidths.append(self.myFont.measure(str(item)))

            #create wors with eqmt data
            for i in self.parent.func_vars.equipment_list:
                self.equipment_tree_rows.append([i.count, i.eqmt_tag, i.path, i.make, i.model, i.sound_level, i.sound_ref_dist, i.tested_q, i.installed_q, i.insertion_loss, i.x_coord, i.y_coord, i.z_coord])

            #getting max widths
            for col_idx in range(len(self.equipment_tree_rows[0])):
                maxWidth = self.maxWidths[col_idx]
                for row in self.equipment_tree_rows:
                    try:
                        currentWidth = self.myFont.measure(str(round(float(row[col_idx]))))
                    except ValueError:
                        currentWidth = self.myFont.measure(str(row[col_idx]))
                    if currentWidth > maxWidth:
                        maxWidth = currentWidth
                self.maxWidths[col_idx] = maxWidth

            #initialize tree
            self.equipment_tree = tkinter.ttk.Treeview(self, columns=self.equipment_tree_columns, show='headings')

            # add rows and colmns to tree
            for col, maxWidth in zip(self.equipment_tree_columns, self.maxWidths):
                self.equipment_tree.heading(col, text=col)
                self.equipment_tree.column(col, minwidth=15, width=maxWidth+25, stretch=0)
            for i, value in enumerate(self.equipment_tree_rows):
                self.equipment_tree.insert("", "end", values=value, tags=self.myFont)
                #sizing
                if i == len(self.equipment_tree_rows)-1:
                    for col in self.equipment_tree_columns:
                        if col in ("eqmt_tag", "model"):
                            width_mult = 10
                            self.equipment_tree.column(col, minwidth=20, width=len(value)*width_mult, stretch=0)

    def generateRcvrTree(self):
        try: # delete tree if already exists
            self.receiver_tree.delete(*self.receiver_tree.get_children())
            self.receiver_tree_rows = []
            for i in self.parent.func_vars.receiver_list:
                self.receiver_tree_rows.append([i.r_name, i.x_coord, i.y_coord, i.z_coord, i.sound_limit, i.predicted_sound_level])
            for i, value in enumerate(self.receiver_tree_rows):
                self.receiver_tree.insert("", "end", values=value, tags=self.myFont)

        except:
            self.receiver_tree_columns = ["R#", "x", "y", "z", "sound limit", "est. level"]
            self.receiver_tree_rows = []
            self.maxWidths = []

            #create widths
            for item in self.receiver_tree_columns:
                self.maxWidths.append(self.myFont.measure(str(item)))

            #create rows with rcvr data
            for i in self.parent.func_vars.receiver_list:
                self.receiver_tree_rows.append([i.r_name, i.x_coord, i.y_coord, i.z_coord, i.sound_limit, i.predicted_sound_level])
            print(self.receiver_tree_rows)

            #getting max widths
            for col_idx in range(len(self.receiver_tree_rows[0])):
                maxWidth = self.maxWidths[col_idx]
                for row in self.receiver_tree_rows:
                    currentWidth = self.myFont.measure(str(row[col_idx]))
                    if currentWidth > maxWidth:
                        maxWidth = currentWidth
                self.maxWidths[col_idx] = maxWidth

            # initializing receiver tree
            self.receiver_tree = tkinter.ttk.Treeview(self, columns=self.receiver_tree_columns, show='headings')

            # adding columns and rows
            for col, maxWidth in zip(self.receiver_tree_columns, self.maxWidths):
                self.receiver_tree.heading(col, text=col)
                self.receiver_tree.column(col, minwidth=15, width=maxWidth+25, stretch=0)
            for i, value in enumerate(self.receiver_tree_rows):
                self.receiver_tree.insert("", "end", values=value, tags=self.myFont)

    def generateBarrierTree(self):
        try: # delete tree if already exists
            self.barrier_tree.delete(*self.barrier_tree.get_children())
            self.barrier_tree_rows = []
            for i in self.parent.func_vars.barrier_list:
                self.barrier_tree_rows.append([i.barrier_name, i.x0_coord, i.y0_coord, i.z0_coord, i.x1_coord, i.y1_coord, i.z1_coord])
            for i, value in enumerate(self.barrier_tree_rows):
                self.barrier_tree.insert("", "end", values=value, tags=self.myFont)

        except:
            self.barrier_tree_columns = ["barrier_name", "x0", "y0", "z0", "x1", "y1", "z1"]
            self.barrier_tree_rows = []
            self.maxWidths = []

            #create widths
            for item in self.barrier_tree_columns:
                self.maxWidths.append(self.myFont.measure(str(item)))

            #create rows with barrier data
            for i in self.parent.func_vars.barrier_list:
                self.barrier_tree_rows.append([i.barrier_name, i.x0_coord, i.y0_coord, i.z0_coord, i.x1_coord, i.y1_coord, i.z1_coord])

            #getting max widths
            for col_idx in range(len(self.barrier_tree_rows[0])):
                maxWidth = self.maxWidths[col_idx]
                for row in self.barrier_tree_rows:
                    currentWidth = self.myFont.measure(str(row[col_idx]))
                    if currentWidth > maxWidth:
                        maxWidth = currentWidth
                self.maxWidths[col_idx] = maxWidth

            # initializing barrier tree
            self.barrier_tree = tkinter.ttk.Treeview(self, columns=self.barrier_tree_columns, show='headings')

            # adding columns and rows
            for col, maxWidth in zip(self.barrier_tree_columns, self.maxWidths):
                self.barrier_tree.heading(col, text=col)
                self.barrier_tree.column(col, minwidth=15, width=maxWidth+25, stretch=0)
            for i, value in enumerate(self.barrier_tree_rows):
                self.barrier_tree.insert("", "end", values=value, tags=self.myFont)


        self.equipment_tree.bind('<ButtonRelease-1>', self.select_item_from_eqmt_tree)
        self.receiver_tree.bind('<ButtonRelease-1>', self.select_item_from_rcvr_tree)
        self.barrier_tree.bind('<ButtonRelease-1>', self.select_item_from_barrier_tree)

    def generateIgnoreMatrixTree(self):
        #todo need to add the eqmt label to the tree
        self.ignore_matrix_tree_columns = ["eqmt"]
        for rcvr in self.parent.func_vars.receiver_list:
            self.ignore_matrix_tree_columns.append(str(rcvr.r_name))
        self.ignore_matrix_tree_rows = self.parent.func_vars.ignore_matrix
        for eqmt, listy in zip(self.parent.func_vars.equipment_list, self.parent.func_vars.ignore_matrix):
            listy.insert(0, eqmt.eqmt_tag)
        self.maxWidths = []

        #create widths
        for item in self.ignore_matrix_tree_columns:
            self.maxWidths.append(self.myFont.measure(str(item)))

        #getting max widths
        for col_idx in range(len(self.ignore_matrix_tree_rows[0])):
            maxWidth = self.maxWidths[col_idx]
            for row in self.ignore_matrix_tree_rows:
                currentWidth = self.myFont.measure(str(row[col_idx]))
                if currentWidth > maxWidth:
                    maxWidth = currentWidth
            self.maxWidths[col_idx] = maxWidth

        # initializing barrier tree
        self.ignore_matrix_tree = tkinter.ttk.Treeview(self, columns=self.ignore_matrix_tree_columns, show='headings')

        # adding columns and rows
        for col in self.ignore_matrix_tree_columns:
            self.ignore_matrix_tree.heading(col, text=col)
            self.ignore_matrix_tree.column(col, minwidth=15, width=maxWidth+25, stretch=0)
        for i, value in enumerate(self.ignore_matrix_tree_rows):
            self.ignore_matrix_tree.insert("", "end", values=value, tags=self.myFont)

    def ARI_interpolation(self, pld, lowerIL, upperIL, lowerPLD, upperPLD):
        diff_in_reduction = (pld-lowerPLD)/(upperPLD-lowerPLD)
        change_IL = upperIL - lowerIL
        barrier_IL = lowerIL + change_IL*diff_in_reduction
        return int(round(barrier_IL,0))

    def ARI_barrier_IL_calc(self, eqmt_x, eqmt_y, eqmt_z, bar_x0, bar_y0, bar_z0, bar_x1, bar_y1, bar_z1, rcvr_x, rcvr_y, rcvr_z):
        #fixing escape on error with same barrier coordinate
        if bar_x0 == bar_x1:
            bar_x0 += 0.0001
            print("corrected bar_x0==bar_x1 error")
        if bar_y0 == bar_y1:
            bar_y0 += 0.0001
            print("corrected bar_y0==bar_y1 error")
        #testing if line of sight is broken along HORIZONTAL plane
        eqmt_point = utils.Point(eqmt_x, eqmt_y)
        receiver_point = utils.Point(rcvr_x, rcvr_y)
        bar_start_point = utils.Point(bar_x0, bar_y0)
        bar_end_point = utils.Point(bar_x1, bar_y1)
        if not utils.doIntersect(eqmt_point, receiver_point, bar_start_point, bar_end_point):
            print("barrier fails horizontal test")
            return 0

        try:
            m_source2receiver = (rcvr_y-eqmt_y)/(rcvr_x-eqmt_x)
        except ZeroDivisionError:
            return 0
        try:
            m_bar_start2end = (bar_y0-bar_y1)/(bar_x0-bar_x1)
        except ZeroDivisionError:
            return 0

        b_source2receiver = eqmt_y-(eqmt_x*m_source2receiver)
        b_bar_start2end = bar_y0-(bar_x0*m_bar_start2end)
        intersection_x = (b_bar_start2end-b_source2receiver)/(m_source2receiver-m_bar_start2end)
        intersection_y = m_source2receiver*intersection_x+b_source2receiver

        bar_min_z = min(bar_z0, bar_z1)
        bar_height_difference = abs(bar_z0-bar_z1)
        bar_length = utils.distance_formula(x0=bar_x0, y0=bar_y0, x1=bar_x1, y1=bar_y1)
        bar_slope = bar_height_difference/bar_length
        if bar_z0 <= bar_z1:
                bar_dist2barxpoint = utils.distance_formula(x0=intersection_x , y0=intersection_y, x1=bar_x0, y1=bar_y0)
        else:
                bar_dist2barxpoint = utils.distance_formula(x0=intersection_x , y0=intersection_y, x1=bar_x1, y1=bar_y1)

        bar_height_to_use = bar_slope*bar_dist2barxpoint+bar_min_z

        # testing if line of sight is broken vertically
        if bar_height_to_use < eqmt_z and bar_height_to_use < rcvr_z:
            print("barrier fails easy vertical test")
            return 0

        distance_source2receiver_horizontal = utils.distance_formula(x0=eqmt_x, y0=eqmt_y, x1=rcvr_x, y1=rcvr_y)
        distance_source2bar_horizontal = utils.distance_formula(x0=eqmt_x, y0=eqmt_y, x1=intersection_x, y1=intersection_y)
        distance_barrier2receiever_straight = distance_source2receiver_horizontal - distance_source2bar_horizontal
        distance_source2receiver_propogation = math.sqrt(distance_source2receiver_horizontal**2+(rcvr_z-eqmt_z)**2)
        distance_source2barrier_top = math.sqrt((bar_height_to_use-eqmt_z)**2+distance_source2bar_horizontal**2)
        distance_receiver2barrier_top = math.sqrt((bar_height_to_use-rcvr_z)**2+distance_barrier2receiever_straight**2)
        path_length_difference = distance_source2barrier_top+distance_receiver2barrier_top-distance_source2receiver_propogation

        #testing if line of sight is broken along VERTICAL plane
        eqmt_point = utils.Point(0, eqmt_z)
        receiver_point = utils.Point(distance_source2receiver_horizontal, rcvr_z)
        bar_start_point = utils.Point(distance_source2bar_horizontal, 0)
        bar_end_point = utils.Point(distance_source2bar_horizontal, bar_height_to_use)
        if not utils.doIntersect(eqmt_point, receiver_point, bar_start_point, bar_end_point):
            print("barrier fails vertical test")
            return 0

        pld = path_length_difference
        if 0 < pld and pld <= 0.5:
            barrier_IL = self.ARI_interpolation(pld, 0, 4, 0, 0.5)
        elif 0.5 < pld and pld <= 1:
            barrier_IL = self.ARI_interpolation(pld, 4, 7, 0.5, 1)
        elif 1 < pld and pld <= 2:
            barrier_IL = self.ARI_interpolation(pld, 7, 10, 1, 2)
        elif 2 < pld and pld <= 3:
            barrier_IL = self.ARI_interpolation(pld, 10, 12, 2, 3)
        elif 3 < pld and pld <= 6:
            barrier_IL = self.ARI_interpolation(pld, 12, 15, 3, 6)
        elif 6 < pld and pld <= 12:
            barrier_IL = self.ARI_interpolation(pld, 15, 17, 6, 12)
        elif 12 < pld:
            barrier_IL = 17
        else:
            barrier_IL = 0

        return [barrier_IL, bar_height_to_use, distance_source2receiver_horizontal, distance_source2bar_horizontal, distance_source2barrier_top, distance_receiver2barrier_top, distance_source2receiver_propogation, path_length_difference, "ARI"]

    def OB_fresnel_barrier_IL_calc(self, eqmt_x, eqmt_y, eqmt_z, hz63, hz125, hz250, hz500, hz1000, hz2000, hz4000, hz8000, eqmt_level, bar_x0, bar_y0, bar_z0, bar_x1, bar_y1, bar_z1, rcvr_x, rcvr_y, rcvr_z):
        #fixing escape on error with same barrier coordinate
        if bar_x0 == bar_x1:
            bar_x0 += 0.0001
            print("corrected bar_x0==bar_x1 error")
        if bar_y0 == bar_y1:
            bar_y0 += 0.0001
            print("corrected bar_y0==bar_y1 error")
        ob_levels_list = [hz63, hz125, hz250, hz500, hz1000, hz2000, hz4000, hz8000]
        ob_bands_list = [63, 125, 250, 500, 1000, 2000, 4000, 8000]
        #testing if line of sight is broken along horizontal plane
        eqmt_point = utils.Point(eqmt_x, eqmt_y)
        receiver_point = utils.Point(rcvr_x, rcvr_y)
        bar_start_point = utils.Point(bar_x0, bar_y0)
        bar_end_point = utils.Point(bar_x1, bar_y1)
        if not utils.doIntersect(eqmt_point, receiver_point, bar_start_point, bar_end_point):
            print("barrier fails horizontal test")
            return 0
        try:
            m_source2receiver = (rcvr_y-eqmt_y)/(rcvr_x-eqmt_x)
        except ZeroDivisionError:
            return 0
        try:
            m_bar_start2end = (bar_y0-bar_y1)/(bar_x0-bar_x1)
        except ZeroDivisionError:
            return 0

        b_source2receiver = eqmt_y-(eqmt_x*m_source2receiver)
        b_bar_start2end = bar_y0-(bar_x0*m_bar_start2end)
        intersection_x = (b_bar_start2end-b_source2receiver)/(m_source2receiver-m_bar_start2end)
        intersection_y = m_source2receiver*intersection_x+b_source2receiver

        bar_min_z = min(bar_z0, bar_z1)
        bar_height_difference = abs(bar_z0-bar_z1)
        bar_length = utils.distance_formula(x0=bar_x0, y0=bar_y0, x1=bar_x1, y1=bar_y1)
        bar_slope = bar_height_difference/bar_length
        if bar_z0 <= bar_z1:
                bar_dist2barxpoint = utils.distance_formula(x0=intersection_x , y0=intersection_y, x1=bar_x0, y1=bar_y0)
        else:
                bar_dist2barxpoint = utils.distance_formula(x0=intersection_x , y0=intersection_y, x1=bar_x1, y1=bar_y1)

        bar_height_to_use = bar_slope*bar_dist2barxpoint+bar_min_z

        # testing if line of sight is broken vertically
        if bar_height_to_use < eqmt_z and bar_height_to_use < rcvr_z:
            print("barrier fails easy vertical test")
            return 0

        distance_source2receiver_horizontal = utils.distance_formula(x0=eqmt_x, y0=eqmt_y, x1=rcvr_x, y1=rcvr_y)
        distance_source2bar_horizontal = utils.distance_formula(x0=eqmt_x, y0=eqmt_y, x1=intersection_x, y1=intersection_y)
        distance_barrier2receiever_straight = distance_source2receiver_horizontal - distance_source2bar_horizontal
        distance_source2receiver_propogation = math.sqrt(distance_source2receiver_horizontal**2+(rcvr_z-eqmt_z)**2)
        distance_source2barrier_top = math.sqrt((bar_height_to_use-eqmt_z)**2+distance_source2bar_horizontal**2)
        distance_receiver2barrier_top = math.sqrt((bar_height_to_use-rcvr_z)**2+distance_barrier2receiever_straight**2)
        path_length_difference = distance_source2barrier_top+distance_receiver2barrier_top-distance_source2receiver_propogation

        #testing if line of sight is broken along VERTICAL plane
        eqmt_point = utils.Point(0, eqmt_z)
        receiver_point = utils.Point(distance_source2receiver_horizontal, rcvr_z)
        bar_start_point = utils.Point(distance_source2bar_horizontal, 0)
        bar_end_point = utils.Point(distance_source2bar_horizontal, bar_height_to_use)
        if not utils.doIntersect(eqmt_point, receiver_point, bar_start_point, bar_end_point):
            print("barrier fails vertical test")
            return 0

        speed_of_sound = 1128
        fresnel_num_list = [(2*path_length_difference)/(speed_of_sound/ob) for ob in ob_bands_list]

        line_point_correction = 0                  #assume no line/point source correction 0 for point, -5 for line
        barrier_finite_infinite_correction = 1.0 #assume infinite barrier see Mehta for correction under finite barrier.
        Kb_barrier_constant = 5                    #assume Kb (barrier constant) for wall = 5, berm = 8
        barrier_attenuate_limit = 20               # wall limit = 20 berm limit = 23

        ob_barrier_attenuation_list = []
        for N in fresnel_num_list:
            n_d = math.sqrt(2*math.pi*N)
            ob_barrier_attenuation = ((20*math.log10(n_d/math.tanh(n_d))) + Kb_barrier_constant + line_point_correction)**barrier_finite_infinite_correction

            if ob_barrier_attenuation > barrier_attenuate_limit:
                ob_barrier_attenuation = barrier_attenuate_limit
            ob_barrier_attenuation_list.append(ob_barrier_attenuation)

        ob_attenuated_levels_list = [x - y for x, y in zip(ob_levels_list, ob_barrier_attenuation_list)]
        ob_a_weighting_list = [-26.2, -16.1, -8.6, -3.2, -0, 1.2, 1, -1.1]
        ob_attenuated_aweighted_levels_list = [x+y for x, y in zip(ob_attenuated_levels_list, ob_a_weighting_list)]

        attenuated_aweighted_level = acoustics.decibel.dbsum(ob_attenuated_aweighted_levels_list)

        barrier_IL = eqmt_level - attenuated_aweighted_level

        return [round(barrier_IL,1), bar_height_to_use, distance_source2receiver_horizontal, distance_source2bar_horizontal, distance_source2barrier_top, distance_receiver2barrier_top, distance_source2receiver_propogation, path_length_difference, "OB-Fresnel"]

    def update_est_noise_levels(self):
        barrierListForExcelOutput_curData = []
        self.barrierListForExcelOutput = [["barrier loss", "eqmt", "rcvr", "bar", "eqmt height", "rcvr height", "bar height", "source to receiver", "source to bar (ft)", "source to bar top", "rcvr to bar top", "direct path", "PLD", "Barrier method", "noise data (if OB Fresnel used)"]]
        for rcvr_index, rcvr in enumerate(self.parent.func_vars.receiver_list):
            print(f"r_name: {rcvr.r_name} x: {rcvr.x_coord}, y: {rcvr.y_coord}, z: {rcvr.z_coord}")
            sound_pressure = 0
            for eqmt_index, eqmt in enumerate(self.parent.func_vars.equipment_list):
                if self.parent.func_vars.ignore_matrix[eqmt_index][rcvr_index] == None:
                    if eqmt.sound_ref_dist == 0:
                        sound_power = eqmt.sound_level + 10*math.log10(eqmt.count)
                    else:
                        q = eqmt.tested_q #need to update this
                        r = eqmt.sound_ref_dist*0.308
                        lp = eqmt.sound_level
                        b = q/(4*math.pi*r**2)
                        sound_power = lp + abs(10*math.log10(b)) + 10*math.log10(eqmt.count)
                    distance = math.sqrt((rcvr.x_coord-eqmt.x_coord)**2 + (rcvr.y_coord - eqmt.y_coord)**2 + (rcvr.z_coord - eqmt.z_coord)**2)
                    try:
                        q = eqmt.installed_q
                        r = distance*0.308
                        attenuation = abs(10*math.log10(q/(4*math.pi*r**2)))
                        used_barrier_name = None
                        barrier_IL = 0
                        if TAKE_ARI_BARRIER == True and TAKE_OB_FRESNAL_BARRIER == False:
                            for bar in self.parent.func_vars.barrier_list:
                                barrier_info_list = self.ARI_barrier_IL_calc(eqmt.x_coord, eqmt.y_coord, eqmt.z_coord, bar.x0_coord, bar.y0_coord, bar.z0_coord, bar.x1_coord, bar.y1_coord, bar.z1_coord, rcvr.x_coord, rcvr.y_coord, rcvr.z_coord)
                                barrier_IL_test = barrier_info_list[0] if barrier_info_list != 0 else 0
                                if barrier_IL_test > barrier_IL:
                                    barrier_IL = barrier_IL_test
                                    used_barrier_name = str(bar.barrier_name + ' - ari')
                                    barrierListForExcelOutput_curData = [barrier_IL, eqmt.eqmt_tag, rcvr.r_name, bar.barrier_name, round(eqmt.z_coord,1), round(rcvr.z_coord, 1), round(barrier_info_list[1],1), round(barrier_info_list[2], 1), round(barrier_info_list[3],1), round(barrier_info_list[4],1), round(barrier_info_list[5],1), round(barrier_info_list[6],1), round(barrier_info_list[7],1), barrier_info_list[8], eqmt.hz63, eqmt.hz125, eqmt.hz250, eqmt.hz500, eqmt.hz1000, eqmt.hz2000, eqmt.hz4000, eqmt.hz8000] if barrier_info_list != 0 else [0]

                        elif TAKE_ARI_BARRIER == True and TAKE_OB_FRESNAL_BARRIER == True:
                            for bar in self.parent.func_vars.barrier_list:
                                if None not in [eqmt.hz63, eqmt.hz125, eqmt.hz250, eqmt.hz500, eqmt.hz1000, eqmt.hz2000, eqmt.hz4000, eqmt.hz8000]:
                                    barrier_info_list = self.OB_fresnel_barrier_IL_calc(eqmt.x_coord, eqmt.y_coord, eqmt.z_coord, eqmt.hz63, eqmt.hz125, eqmt.hz250, eqmt.hz500, eqmt.hz1000, eqmt.hz2000, eqmt.hz4000, eqmt.hz8000, eqmt.sound_level, bar.x0_coord, bar.y0_coord, bar.z0_coord, bar.x1_coord, bar.y1_coord, bar.z1_coord, rcvr.x_coord, rcvr.y_coord, rcvr.z_coord)
                                    barrier_IL_test = barrier_info_list[0] if barrier_info_list != 0 else 0
                                    barriermethod = ' - OB_fresnel'
                                else:
                                    barrier_info_list = self.ARI_barrier_IL_calc(eqmt.x_coord, eqmt.y_coord, eqmt.z_coord, bar.x0_coord, bar.y0_coord, bar.z0_coord, bar.x1_coord, bar.y1_coord, bar.z1_coord, rcvr.x_coord, rcvr.y_coord, rcvr.z_coord)
                                    barrier_IL_test = barrier_info_list[0] if barrier_info_list != 0 else 0
                                    barriermethod = ' - ari'
                                if barrier_IL_test > barrier_IL:
                                    barrier_IL = barrier_IL_test
                                    used_barrier_name = str(bar.barrier_name + barriermethod)
                                    barrierListForExcelOutput_curData = [int(round(barrier_IL,0)), eqmt.eqmt_tag, rcvr.r_name, bar.barrier_name, round(eqmt.z_coord,1), round(rcvr.z_coord, 1), round(barrier_info_list[1],1), round(barrier_info_list[2], 1), round(barrier_info_list[3],1), round(barrier_info_list[4],1), round(barrier_info_list[5],1), round(barrier_info_list[6],1), round(barrier_info_list[7],1), barrier_info_list[8], eqmt.hz63, eqmt.hz125, eqmt.hz250, eqmt.hz500, eqmt.hz1000, eqmt.hz2000, eqmt.hz4000, eqmt.hz8000] if barrier_info_list != 0 else [0]
                        try:
                            self.barrierListForExcelOutput.append(barrierListForExcelOutput_curData)
                        except UnboundLocalError:
                            print("Barrier Calculation Block Error")

                        barrierListForExcelOutput_curData = []
                        # print(eqmt.eqmt_tag, " - ", barrier_IL, int(barrier_IL), int(round(barrier_IL, 0)))
                        spl = sound_power-eqmt.insertion_loss-attenuation-barrier_IL
                        # if barriermethod == ' - OB_fresnel':
                        print(f"eqmt: __{eqmt.eqmt_tag}, rcvr: __{rcvr.r_name}, bar: __{used_barrier_name}, barrier IL: __{barrier_IL}")

                    except (ValueError, ZeroDivisionError):
                        print('MATH DOMAIN ERROR OCCURED')
                        spl = 1000

                elif self.parent.func_vars.ignore_matrix[eqmt_index][rcvr_index] != None:
                    self.barrierListForExcelOutput.append(barrierListForExcelOutput_curData)
                    spl = 0
                sound_pressure += 10**(spl/10)
                # print(f"eqmt, x: {eqmt.x_coord}, y: {eqmt.y_coord}, z: {eqmt.z_coord}, lwa: {round(sound_power,0)}, IL: {round(eqmt.insertion_loss,0)}, distance: {round(distance,1)}, attenuation: {round(attenuation,1)}")
            rcvr.predicted_sound_level = round(10*math.log10(sound_pressure),1)
            # if rcvr.r_name == "R1":
            #     print(f"predicted sound level: {rcvr.predicted_sound_level}")
                # print(f"distance: {distance}")
            for listy in self.barrierListForExcelOutput:
                print(listy, "/n")

    def select_item_from_eqmt_tree(self, event):
        self.deselect_item_from_trees()
        self.current_equipment = self.equipment_tree.focus()
        self.current_equipment = self.equipment_tree.item(self.current_equipment)['values']
        print(self.current_equipment)

    def select_item_from_rcvr_tree(self, event):
        self.deselect_item_from_trees()
        self.current_receiver = self.receiver_tree.focus()
        self.current_receiver = self.receiver_tree.item(self.current_receiver)['values']
        print(self.current_receiver)

    def select_item_from_barrier_tree(self, event):
        self.deselect_item_from_trees()
        self.current_barrier = self.barrier_tree.focus()
        self.current_barrier = self.barrier_tree.item(self.current_barrier)['values']
        print(self.current_barrier)

    def deselect_item_from_trees(self):
        self.current_barrier = None
        self.current_receiver = None
        self.current_equipment = None

    def onExportListButton(self):
        wb = openpyxl.load_workbook(XL_FILEPATH, keep_vba=True, data_only=False)
        ws = wb['Input LwA_XYZ']

        for obj in self.parent.func_vars.equipment_list:
            for row in ws.iter_rows(max_row=100):
                if row[1].value == None:
                    break
                if row[1].value.replace(" ","-") == obj.eqmt_tag.replace(" ", "-"):
                    row[19].value = obj.x_coord
                    row[20].value = obj.y_coord

        for obj in self.parent.func_vars.receiver_list:
            for row in ws.iter_rows(max_row=100):
                if row[25].value == None:
                    break
                if row[25].value.replace(" ","-") == obj.r_name.replace(" ", "-"):
                    row[26].value = obj.x_coord
                    row[27].value = obj.y_coord

        for obj in self.parent.func_vars.barrier_list:
            for row in ws.iter_rows(min_row=24, max_row=100):
                if row[25].value == None:
                    break
                if row[25].value.replace(" ","-") == obj.barrier_name.replace(" ", "-"):
                    row[26].value = obj.x0_coord
                    row[27].value = obj.y0_coord
                    row[28].value = obj.z0_coord
                    row[29].value = obj.x1_coord
                    row[30].value = obj.y1_coord
                    row[31].value = obj.z1_coord

        barCalcListNum = 1
        totalEqmtCount = len(self.parent.func_vars.equipment_list)
        print(totalEqmtCount)
        for col in range(73, 87):
            for row in ws.iter_rows(min_row=2, max_row=2+totalEqmtCount-1):
                print(col)
                if barCalcListNum > len(self.barrierListForExcelOutput)-1:
                    break
                if not self.barrierListForExcelOutput[barCalcListNum]:
                    row[col].value = 0
                    print(0)
                else:
                    row[col].value = self.barrierListForExcelOutput[barCalcListNum][0]
                    print(self.barrierListForExcelOutput[barCalcListNum][0])

                barCalcListNum+=1

        # saving scale
        ws['AE20'] = self.parent.func_vars.known_distance_ft
        ws['AF20'] = self.parent.func_vars.scale_line_distance_px

        print("saving")
        wb.save(filename=XL_FILEPATH_SAVE)
        print("saved")
        # wb.close()

    def e1_unfocus(self, event):
        self.status_label.focus()

    def e1_select_all(self, event):
        e1_text = self.e1.get()
        self.e1.selection_range(0, len(e1_text))

    def save_changes(self):
        offset = 20
        if self.current_equipment:
            # self, count, eqmt_tag, path, make, model, sound_level, sound_ref_dist, tested_q, installed_q, insertion_loss, x_coord, y_coord, z_coord
            self.current_obj.count = float(self.count_input.get())
            self.current_obj.eqmt_tag = self.eqmt_tag_input.get()
            self.current_obj.path = self.path_input.get()
            self.current_obj.make = self.make_input.get()
            self.current_obj.model = self.model_input.get()
            self.current_obj.sound_level = float(self.sound_level_input.get())
            self.current_obj.sound_ref_dist = float(self.sound_ref_dist_input.get())
            self.current_obj.tested_q = float(self.tested_q_input.get())
            self.current_obj.installed_q = float(self.installed_q_input.get())
            self.current_obj.insertion_loss = float(self.insertion_loss_input.get())
            self.current_obj.x_coord = float(self.x_coord_input.get())
            self.current_obj.y_coord = float(self.y_coord_input.get())
            self.current_obj.z_coord = float(self.z_coord_input.get())

            self.eqmt_tagged = self.parent.editor.canvas.gettags(self.current_obj.eqmt_tag)
            self.eqmt_num = self.eqmt_tagged[0]
            self.eqmt_ids = self.parent.editor.canvas.find_withtag(self.current_obj.eqmt_tag)
            self.current_shape = self.eqmt_ids[0]
            self.current_text = self.eqmt_ids[1]

            x = self.current_obj.x_coord/self.parent.func_vars.master_scale
            y = self.current_obj.y_coord/self.parent.func_vars.master_scale
            self.parent.editor.canvas.coords(self.current_shape, x+offset, y+offset, x-offset, y-offset)
            self.parent.editor.canvas.coords(self.current_text, x, y)

        if self.current_receiver:
            # self, r_name, x_coord, y_coord, z_coord, sound_limit, predicted_sound_level
            self.current_obj.r_name = self.r_name_input.get()
            self.current_obj.x_coord = float(self.x_coord_input.get())
            self.current_obj.y_coord = float(self.y_coord_input.get())
            self.current_obj.z_coord = float(self.z_coord_input.get())
            self.current_obj.sound_limit = float(self.sound_limit_input.get())

            self.rcvr_tagged = self.parent.editor.canvas.gettags(self.current_obj.r_name)
            self.rcvr_num = self.rcvr_tagged[0]
            self.rcvr_ids = self.parent.editor.canvas.find_withtag(self.current_obj.r_name)
            self.current_shape = self.rcvr_ids[0]
            self.current_text = self.rcvr_ids[1]

            x = self.current_obj.x_coord/self.parent.func_vars.master_scale
            y = self.current_obj.y_coord/self.parent.func_vars.master_scale
            self.parent.editor.canvas.coords(self.current_shape, x+offset, y+offset, x-offset, y-offset)
            self.parent.editor.canvas.coords(self.current_text, x, y)

        if self.current_barrier:
            # self, barrier_name, x0_coord, y0_coord, z0_coord, x1_coord, y1_coord, z1_coord
            self.current_obj.barrier_name = self.barrier_name_input.get()
            self.current_obj.x0_coord = float(self.x0_coord_input.get())
            self.current_obj.y0_coord = float(self.y0_coord_input.get())
            self.current_obj.z0_coord = float(self.z0_coord_input.get())
            self.current_obj.x1_coord = float(self.x1_coord_input.get())
            self.current_obj.y1_coord = float(self.y1_coord_input.get())
            self.current_obj.z1_coord = float(self.z1_coord_input.get())

            self.barr_tagged = self.parent.editor.canvas.gettags(self.current_obj.barrier_name)
            self.barr_num = self.barr_tagged[0]
            self.barr_ids = self.parent.editor.canvas.find_withtag(self.current_obj.barrier_name)
            self.current_shape = self.barr_ids[0]
            self.current_text = self.barr_ids[1]

            print(self.current_obj.x0_coord)
            print(self.current_obj.y0_coord)
            print(self.current_obj.x1_coord)
            print(self.current_obj.y1_coord)

            x0 = self.current_obj.x0_coord/self.parent.func_vars.master_scale
            y0 = self.current_obj.y0_coord/self.parent.func_vars.master_scale
            x1 = self.current_obj.x1_coord/self.parent.func_vars.master_scale
            y1 = self.current_obj.y1_coord/self.parent.func_vars.master_scale

            self.parent.editor.canvas.coords(self.current_shape, x0, y0, x1, y1)
            self.parent.editor.canvas.coords(self.current_text, x0 + (x1-x0)/2, y0 + (y1 - y0)/2)
            print("Hey", 2.85/self.parent.func_vars.master_scale)



        self.update_est_noise_levels()
        self.generateEqmtTree()
        self.generateRcvrTree()
        self.generateBarrierTree()
        self.newWindow.destroy()

    def open_item_editor_window(self, event):
        self.newWindow = tkinter.Toplevel()
        self.newWindow.title("item editor")
        self.newWindow.geometry("500x500")

        if self.current_equipment:
            # self, count, eqmt_tag, path, make, model, sound_level, sound_ref_dist, tested_q, installed_q, insertion_loss, x_coord, y_coord, z_coord
            for obj in self.parent.func_vars.equipment_list:
                if obj.eqmt_tag == self.current_equipment[1]:
                    self.current_obj = obj
                    break

            self.count_label = tkinter.Label(self.newWindow, text="count", borderwidth=2, font=(None, 15))
            self.eqmt_tag_label = tkinter.Label(self.newWindow, text="eqmt_tag", borderwidth=2, font=(None, 15))
            self.path_label = tkinter.Label(self.newWindow, text="path", borderwidth=2, font=(None, 15))
            self.make_label = tkinter.Label(self.newWindow, text="make", borderwidth=2, font=(None, 15))
            self.model_label = tkinter.Label(self.newWindow, text="model", borderwidth=2, font=(None, 15))
            self.sound_level_label = tkinter.Label(self.newWindow, text="sound_level", borderwidth=2, font=(None, 15))
            self.sound_ref_dist_label = tkinter.Label(self.newWindow, text="sound_ref_dist", borderwidth=2, font=(None, 15))
            self.tested_q_label = tkinter.Label(self.newWindow, text="tested_q", borderwidth=2, font=(None, 15))
            self.installed_q_label = tkinter.Label(self.newWindow, text="installed_q", borderwidth=2, font=(None, 15))
            self.insertion_loss_label = tkinter.Label(self.newWindow, text="insertion_loss", borderwidth=2, font=(None, 15))
            self.x_coord_label = tkinter.Label(self.newWindow, text="x_coord", borderwidth=2, font=(None, 15))
            self.y_coord_label = tkinter.Label(self.newWindow, text="y_coord", borderwidth=2, font=(None, 15))
            self.z_coord_label = tkinter.Label(self.newWindow, text="z_coord", borderwidth=2, font=(None, 15))

            self.count_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.eqmt_tag_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.path_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.make_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.model_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.sound_level_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.sound_ref_dist_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.tested_q_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.installed_q_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.insertion_loss_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.x_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.y_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.z_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)

            self.count_label.grid(row=0, column=0, sticky=tkinter.N + tkinter.W)
            self.eqmt_tag_label.grid(row=1, column=0, sticky=tkinter.N + tkinter.W)
            self.path_label.grid(row=2, column=0, sticky=tkinter.N + tkinter.W)
            self.make_label.grid(row=3, column=0, sticky=tkinter.N + tkinter.W)
            self.model_label.grid(row=4, column=0, sticky=tkinter.N + tkinter.W)
            self.sound_level_label.grid(row=5, column=0, sticky=tkinter.N + tkinter.W)
            self.sound_ref_dist_label.grid(row=6, column=0, sticky=tkinter.N + tkinter.W)
            self.tested_q_label.grid(row=7, column=0, sticky=tkinter.N + tkinter.W)
            self.installed_q_label.grid(row=8, column=0, sticky=tkinter.N + tkinter.W)
            self.insertion_loss_label.grid(row=9, column=0, sticky=tkinter.N + tkinter.W)
            self.x_coord_label.grid(row=10, column=0, sticky=tkinter.N + tkinter.W)
            self.y_coord_label.grid(row=11, column=0, sticky=tkinter.N + tkinter.W)
            self.z_coord_label.grid(row=12, column=0, sticky=tkinter.N + tkinter.W)

            self.count_input.grid(row=0, column=1, sticky=tkinter.N + tkinter.W)
            self.eqmt_tag_input.grid(row=1, column=1, sticky=tkinter.N + tkinter.W)
            self.path_input.grid(row=2, column=1, sticky=tkinter.N + tkinter.W)
            self.make_input.grid(row=3, column=1, sticky=tkinter.N + tkinter.W)
            self.model_input.grid(row=4, column=1, sticky=tkinter.N + tkinter.W)
            self.sound_level_input.grid(row=5, column=1, sticky=tkinter.N + tkinter.W)
            self.sound_ref_dist_input.grid(row=6, column=1, sticky=tkinter.N + tkinter.W)
            self.tested_q_input.grid(row=7, column=1, sticky=tkinter.N + tkinter.W)
            self.installed_q_input.grid(row=8, column=1, sticky=tkinter.N + tkinter.W)
            self.insertion_loss_input.grid(row=9, column=1, sticky=tkinter.N + tkinter.W)
            self.x_coord_input.grid(row=10, column=1, sticky=tkinter.N + tkinter.W)
            self.y_coord_input.grid(row=11, column=1, sticky=tkinter.N + tkinter.W)
            self.z_coord_input.grid(row=12, column=1, sticky=tkinter.N + tkinter.W)

            self.count_input.insert(0, self.current_obj.count)
            self.eqmt_tag_input.insert(0, self.current_obj.eqmt_tag)
            self.path_input.insert(0, self.current_obj.path)
            self.make_input.insert(0, self.current_obj.make)
            self.model_input.insert(0, self.current_obj.model)
            self.sound_level_input.insert(0, self.current_obj.sound_level)
            self.sound_ref_dist_input.insert(0, self.current_obj.sound_ref_dist)
            self.tested_q_input.insert(0, self.current_obj.tested_q)
            self.installed_q_input.insert(0, self.current_obj.installed_q)
            self.insertion_loss_input.insert(0, self.current_obj.insertion_loss)
            self.x_coord_input.insert(0, self.current_obj.x_coord)
            self.y_coord_input.insert(0, self.current_obj.y_coord)
            self.z_coord_input.insert(0, self.current_obj.z_coord)

        if self.current_receiver:
            # self, r_name, x_coord, y_coord, z_coord, sound_limit, predicted_sound_level
            for obj in self.parent.func_vars.receiver_list:
                if obj.r_name == self.current_receiver[0]:
                    self.current_obj = obj
                    break

            self.r_name_label = tkinter.Label(self.newWindow, text="r_name", borderwidth=2, font=(None, 15))
            self.x_coord_label = tkinter.Label(self.newWindow, text="x_coord", borderwidth=2, font=(None, 15))
            self.y_coord_label = tkinter.Label(self.newWindow, text="y_coord", borderwidth=2, font=(None, 15))
            self.z_coord_label = tkinter.Label(self.newWindow, text="z_coord", borderwidth=2, font=(None, 15))
            self.sound_limit_label = tkinter.Label(self.newWindow, text="sound_limit", borderwidth=2, font=(None, 15))

            self.r_name_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.x_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.y_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.z_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.sound_limit_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)

            self.r_name_label.grid(row=0, column=0, sticky=tkinter.N + tkinter.W)
            self.x_coord_label.grid(row=1, column=0, sticky=tkinter.N + tkinter.W)
            self.y_coord_label.grid(row=2, column=0, sticky=tkinter.N + tkinter.W)
            self.z_coord_label.grid(row=3, column=0, sticky=tkinter.N + tkinter.W)
            self.sound_limit_label.grid(row=4, column=0, sticky=tkinter.N + tkinter.W)

            self.r_name_input.grid(row=0, column=1, sticky=tkinter.N + tkinter.W)
            self.x_coord_input.grid(row=1, column=1, sticky=tkinter.N + tkinter.W)
            self.y_coord_input.grid(row=2, column=1, sticky=tkinter.N + tkinter.W)
            self.z_coord_input.grid(row=3, column=1, sticky=tkinter.N + tkinter.W)
            self.sound_limit_input.grid(row=4, column=1, sticky=tkinter.N + tkinter.W)

            self.r_name_input.insert(0, self.current_obj.r_name)
            self.x_coord_input.insert(0, self.current_obj.x_coord)
            self.y_coord_input.insert(0, self.current_obj.y_coord)
            self.z_coord_input.insert(0, self.current_obj.z_coord)
            self.sound_limit_input.insert(0, self.current_obj.sound_limit)

        if self.current_barrier:
            # self, barrier_name, x0_coord, y0_coord, z0_coord, x1_coord, y1_coord, z1_coord
            for obj in self.parent.func_vars.barrier_list:
                if obj.barrier_name == self.current_barrier[0]:
                    self.current_obj = obj
                    break

            self.barrier_name_label = tkinter.Label(self.newWindow, text="barrier_name", borderwidth=2, font=(None, 15))
            self.x0_coord_label = tkinter.Label(self.newWindow, text="x0_coord", borderwidth=2, font=(None, 15))
            self.y0_coord_label = tkinter.Label(self.newWindow, text="y0_coord", borderwidth=2, font=(None, 15))
            self.z0_coord_label = tkinter.Label(self.newWindow, text="z0_coord", borderwidth=2, font=(None, 15))
            self.x1_coord_label = tkinter.Label(self.newWindow, text="x1_coord", borderwidth=2, font=(None, 15))
            self.y1_coord_label = tkinter.Label(self.newWindow, text="y1_coord", borderwidth=2, font=(None, 15))
            self.z1_coord_label = tkinter.Label(self.newWindow, text="z1_coord", borderwidth=2, font=(None, 15))

            self.barrier_name_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.x0_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.y0_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.z0_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.x1_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.y1_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)
            self.z1_coord_input = tkinter.Entry(self.newWindow, font=(None, 15), width=24)

            self.barrier_name_label.grid(row=0, column=0, sticky=tkinter.N + tkinter.W)
            self.x0_coord_label.grid(row=1, column=0, sticky=tkinter.N + tkinter.W)
            self.y0_coord_label.grid(row=2, column=0, sticky=tkinter.N + tkinter.W)
            self.z0_coord_label.grid(row=3, column=0, sticky=tkinter.N + tkinter.W)
            self.x1_coord_label.grid(row=4, column=0, sticky=tkinter.N + tkinter.W)
            self.y1_coord_label.grid(row=5, column=0, sticky=tkinter.N + tkinter.W)
            self.z1_coord_label.grid(row=6, column=0, sticky=tkinter.N + tkinter.W)

            self.barrier_name_input.grid(row=0, column=1, sticky=tkinter.N + tkinter.W)
            self.x0_coord_input.grid(row=1, column=1, sticky=tkinter.N + tkinter.W)
            self.y0_coord_input.grid(row=2, column=1, sticky=tkinter.N + tkinter.W)
            self.z0_coord_input.grid(row=3, column=1, sticky=tkinter.N + tkinter.W)
            self.x1_coord_input.grid(row=4, column=1, sticky=tkinter.N + tkinter.W)
            self.y1_coord_input.grid(row=5, column=1, sticky=tkinter.N + tkinter.W)
            self.z1_coord_input.grid(row=6, column=1, sticky=tkinter.N + tkinter.W)

            self.barrier_name_input.insert(0, self.current_obj.barrier_name)
            self.x0_coord_input.insert(0, self.current_obj.x0_coord)
            self.y0_coord_input.insert(0, self.current_obj.y0_coord)
            self.z0_coord_input.insert(0, self.current_obj.z0_coord)
            self.x1_coord_input.insert(0, self.current_obj.x1_coord)
            self.y1_coord_input.insert(0, self.current_obj.y1_coord)
            self.z1_coord_input.insert(0, self.current_obj.z1_coord)

        self.save_changes_button = tkinter.Button(self.newWindow, text="Save Changes", command=self.save_changes, font=(None, 15))
        self.save_changes_button.grid(row=15, column=1, columnspan=2, sticky=tkinter.N)

class Main_Application(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self) # , parent
        self.parent = parent

        self.func_vars = FuncVars(self)
        self.editor = Editor(self)
        self.pane_toolbox = Pane_Toolbox(self)
        self.pane_eqmt_info = Pane_Eqmt_Info(self)

        self.editor.grid(row=0, rowspan=2, column=0, stick=tkinter.N)
        self.pane_toolbox.grid(row=0, column=1, padx=20, pady=20, stick=tkinter.N+tkinter.W)
        self.pane_eqmt_info.grid(row=1, column=1, padx=20, pady=20, stick=tkinter.N)

def main():
    root = tkinter.Tk()
    mainApp = Main_Application(root)
    mainApp.pack(side="top", fill="both", expand=True)
    root.geometry('+0+0') #puts window in top left
    root.mainloop()


if __name__ == "__main__":
    main()
